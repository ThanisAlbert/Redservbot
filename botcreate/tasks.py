from datetime import time, date


from celery import shared_task
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl.workbook import Workbook

from Redservbot import settings

@shared_task(bind=True)
def mailsupport(self):
    import logging
    logger = logging.getLogger(__name__)
    logger.info("Testing Mailsupport")

@shared_task(bind=True)
def mail_admin(self):
    from datetime import datetime
    import time
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    import smtplib
    from botcreate.models import Bot
    import logging
    from selenium.webdriver.chrome.options import Options
    from selenium import webdriver
    from email.mime.application import MIMEApplication
    from email.mime.image import MIMEImage

    logger = logging.getLogger(__name__)

    message = "Please find the attached Daily Automation Report (01-July-2023 to 30-September-2023 - Q2)"

    logger.info("Testing")

    # =========================================================ChartPreparation======================================

    options = Options()
    options.headless = False  # Run Chrome in headless mode

    url = "http://172.26.1.19:85/bot/mailchart.html"
    #url = "http://127.0.0.1:8000/bot/mailchart.html"

    driver = webdriver.Chrome(executable_path="E:\\chromedriver.exe", options=options)
    # driver = webdriver.Chrome(executable_path="D:\\chromedriver.exe", options=options)

    driver.get(url)

    # Wait for the chart to be generated (replace "chart-element-id" with the actual ID of the chart element)
    chart_element = driver.find_element_by_id("mailchartcontainer")

    time.sleep(3)

    chart_image = chart_element.screenshot_as_png

    with open(settings.MEDIA_ROOT + "\\chart.png", "wb") as f:
        f.write(chart_image)

        # ===================================Chart2==========================================

        url = "http://172.26.1.19:85/bot/mailchart2.html"
        #url = "http://127.0.0.1:8000/bot/mailchart2.html"

        driver = webdriver.Chrome(executable_path="E:\\chromedriver.exe", options=options)
        # driver = webdriver.Chrome(executable_path="D:\\chromedriver.exe", options=options)

        driver.get(url)

        # Wait for the chart to be generated (replace "chart-element-id" with the actual ID of the chart element)
        chart_element = driver.find_element_by_id("mailchartcontainer")

        time.sleep(3)

        chart_image = chart_element.screenshot_as_png

        with open(settings.MEDIA_ROOT + "\\chart2.png", "wb") as f:
            f.write(chart_image)

    # ==============================================================Q4 Starts======================================================
    from_date_q4 = "2024-01-01"
    to_date_q4 = "2024-03-31"

    New_Bots_q4 = Bot.objects.filter(
        Q(Startdate__range=(from_date_q4, to_date_q4)) | Q(Enddate__range=(from_date_q4, to_date_q4)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots_q4 = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date_q4, to_date_q4)) | Q(enhanceenddate__range=(from_date_q4, to_date_q4)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots_q4 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q4, to_date_q4)) | Q(Startdate__range=(from_date_q4, to_date_q4)) | Q(
            Enddate__range=(from_date_q4, to_date_q4)) | Q(enhancestartdate__range=(from_date_q4, to_date_q4)) | Q(
            enhanceenddate__range=(from_date_q4, to_date_q4)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports_q4 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q4, to_date_q4)) | Q(Startdate__range=(from_date_q4, to_date_q4)) | Q(
            Enddate__range=(from_date_q4, to_date_q4)) | Q(enhancestartdate__range=(from_date_q4, to_date_q4)) | Q(
            enhanceenddate__range=(from_date_q4, to_date_q4))).order_by('Developermail')

    table_row_data_q4 = ""
    enhance_table_row_data_q4 = ""
    rest_row_data_q4 = ""

    if New_Bots_q4.exists():

        for obj in New_Bots_q4:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data_q4 = table_row_data_q4 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data_q4 = "No_data"

    if Enh_Bots_q4.exists():

        for obj in Enh_Bots_q4:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data_q4 = enhance_table_row_data_q4 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data_q4 = "No_data"

    if rest_Bots_q4.exists():

        for obj in rest_Bots_q4:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data_q4 = rest_row_data_q4 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data_q4 = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports_q4:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q4_Jan_Mar_2024.xlsx")

    # ==============================================================q4 Ends========================================================

    # =============================================================Q3 Starts======================================================

    from_date_q3 = "2023-10-01"
    to_date_q3 = "2023-12-31"

    New_Bots_q3 = Bot.objects.filter(
        Q(Startdate__range=(from_date_q3, to_date_q3)) | Q(Enddate__range=(from_date_q3, to_date_q3)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots_q3 = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date_q3, to_date_q3)) | Q(enhanceenddate__range=(from_date_q3, to_date_q3)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots_q3 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q3, to_date_q3)) | Q(Startdate__range=(from_date_q3, to_date_q3)) | Q(
            Enddate__range=(from_date_q3, to_date_q3)) | Q(enhancestartdate__range=(from_date_q3, to_date_q3)) | Q(
            enhanceenddate__range=(from_date_q3, to_date_q3)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports_q3 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q3, to_date_q3)) | Q(Startdate__range=(from_date_q3, to_date_q3)) | Q(
            Enddate__range=(from_date_q3, to_date_q3)) | Q(enhancestartdate__range=(from_date_q3, to_date_q3)) | Q(
            enhanceenddate__range=(from_date_q3, to_date_q3))).order_by('Developermail')

    table_row_data_q3 = ""
    enhance_table_row_data_q3 = ""
    rest_row_data_q3 = ""

    if New_Bots_q3.exists():

        for obj in New_Bots_q3:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data_q3 = table_row_data_q3 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data_q3 = "No_data"

    if Enh_Bots_q3.exists():

        for obj in Enh_Bots_q3:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data_q3 = enhance_table_row_data_q3 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data_q3 = "No_data"

    if rest_Bots_q3.exists():

        for obj in rest_Bots_q3:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data_q3 = rest_row_data_q3 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data_q3 = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports_q3:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q3_Oct_Dec_2023.xlsx")

    # =============================================================Q3 Ends=========================================================

    # ==============================================================Q2 Starts======================================================

    from_date = "2023-07-01"
    to_date = "2023-09-30"

    New_Bots = Bot.objects.filter(Q(Startdate__range=(from_date, to_date)) | Q(Enddate__range=(from_date, to_date)),
                                  Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date, to_date)) | Q(enhanceenddate__range=(from_date, to_date)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots = Bot.objects.filter(
        Q(Creationdate__range=(from_date, to_date)) | Q(Startdate__range=(from_date, to_date)) | Q(
            Enddate__range=(from_date, to_date)) | Q(enhancestartdate__range=(from_date, to_date)) | Q(
            enhanceenddate__range=(from_date, to_date)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports = Bot.objects.filter(
        Q(Creationdate__range=(from_date, to_date)) | Q(Startdate__range=(from_date, to_date)) | Q(
            Enddate__range=(from_date, to_date)) | Q(enhancestartdate__range=(from_date, to_date)) | Q(
            enhanceenddate__range=(from_date, to_date))).order_by('Developermail')

    table_row_data = ""
    enhance_table_row_data = ""
    rest_row_data = ""

    if New_Bots.exists():

        for obj in New_Bots:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data = table_row_data + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data = "No_data"

    if Enh_Bots.exists():

        for obj in Enh_Bots:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data = enhance_table_row_data + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data = "No_data"

    if rest_Bots.exists():

        for obj in rest_Bots:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data = rest_row_data + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q2_July_Sept_2023.xlsx")

    # ==========================================================q2 ends=================================================

    #===========================================================q1 starts===============================================


    from_date_q1 = "2023-04-01"
    to_date_q1 = "2023-06-30"

    New_Bots_q1 = Bot.objects.filter(
        Q(Startdate__range=(from_date_q1, to_date_q1)) | Q(Enddate__range=(from_date_q1, to_date_q1)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots_q1 = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date_q1, to_date_q1)) | Q(enhanceenddate__range=(from_date_q1, to_date_q1)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots_q1 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q1, to_date_q1)) | Q(Startdate__range=(from_date_q1, to_date_q1)) | Q(
            Enddate__range=(from_date_q1, to_date_q1)) | Q(enhancestartdate__range=(from_date_q1, to_date_q1)) | Q(
            enhanceenddate__range=(from_date_q1, to_date_q1)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports_q1 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q1, to_date_q1)) | Q(Startdate__range=(from_date_q1, to_date_q1)) | Q(
            Enddate__range=(from_date_q1, to_date_q1)) | Q(enhancestartdate__range=(from_date_q1, to_date_q1)) | Q(
            enhanceenddate__range=(from_date_q1, to_date_q1))).order_by('Developermail')

    table_row_data_q1 = ""
    enhance_table_row_data_q1 = ""
    rest_row_data_q1 = ""

    if New_Bots_q1.exists():

        for obj in New_Bots_q1:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data_q1 = table_row_data_q1 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data_q1 = "No_data"

    if Enh_Bots_q1.exists():

        for obj in Enh_Bots_q1:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data_q1 = enhance_table_row_data_q1 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data_q1 = "No_data"

    if rest_Bots_q1.exists():

        for obj in rest_Bots_q1:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data_q1 = rest_row_data_q1 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data_q1 = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports_q1:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q1_Apr_Jun_2023.xlsx")


    #============================================================q1 ends=================================================



    done = 0

    while done == 0:
        try:
            msg = MIMEMultipart()
            msg['From'] = 'botomation@redingtongroup.com'

            to_addr = []
            #to_addr.append("prajnya.sahu@redingtongroup.com")
            #to_addr.append("prasanth.muruga@redingtongroup.com")
            #to_addr.append("santhosh.arjunan@redingtongroup.com")
            to_addr.append("thanis.albert@redingtongroup.com")

            cc_addr = []
            #cc_addr.append("rathina.moorthy@redingtongroup.com")
            #cc_addr.append("raheema.shahul@redingtongroup.com")
            #cc_addr.append("srinivasa.babu@redingtongroup.com")

            bc_addr = []
            bc_addr.append("thanis.a@redingtongroup.com")

            msg['To'] = ', '.join(to_addr)
            msg['Cc'] = ', '.join(cc_addr)
            msg['Bcc'] = ', '.join(bc_addr)

            recipients = to_addr + bc_addr + cc_addr

            msg['Subject'] = 'Daily Automation Report (01-Jan-2024 to 31-March-2024 - Q4)'

            html = '''
                                                                       <html>
                                                                       <style>
                                                                         table,tr,td {
                                                                         border: 1px solid black;
                                                                         border-collapse:collapse;                          
                                                                         }
                                                                         span{
                                                                         color:red;
                                                                         }
                                                                       </style>
                                                                       <body>                                                                   

                                                                       Dear Team,
                                                                       <br><br>

                                                                       <br>  

                                                                       <img src="cid:image2">

                                                                       <br><br><br><br><br>

                                                                       <img src="cid:image1">

                                                                       <br><br><br><br>
                                                                       Please find the attached Daily Automation Report (01-January-2024 to 31-March-2024 - Q4)
                                                                       <br> 


                                                                       <br><br>*************************************************************************<b style="color:green;">(01-January-2024 to 31-March-2024 - Q4)</b>******************************************************<br><br><br>

                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data_q4 + '''                                             
                                                                       </table> 

                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data_q4 + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data_q4 + '''                                             
                                                                       </table>


                                                                       <br><br>*************************************************************************<b style="color:green;">(01-October-2023 to 31-December-2023 - Q3)</b>******************************************************<br><br><br>

                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data_q3 + '''                                             
                                                                       </table> 

                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data_q3 + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data_q3 + '''                                             
                                                                       </table>



                                                                        <br><br><br><br><br>*************************************************************************<b style="color:green;">(01-July-2023 to 30-September-2023 - Q2)</b>********************************************************<br><br><br><br>


                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data + '''                                             
                                                                       </table>                                                                   


                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data + '''                                             
                                                                       </table>
                                                                       
                                                                       <br><br>
                                                                       <br><br>*************************************************************************<b style="color:green;">(01-April-2023 to 31-June-2023 - Q1)</b>******************************************************<br><br><br>

                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data_q1 + '''                                             
                                                                       </table> 

                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data_q1 + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data_q1 + '''                                             
                                                                       </table>




                                                                       <br>                                                           
                                                               Bot Application Link: <a href ="http://172.26.1.19:85/bot/">http://172.26.1.19:85/bot/</a>                                               
                                                               <br><br>This is an automatically generated email. <br><br>
                                                                       Thanks & Regards,<br>
                                                                       Automation Team - RGS
                                                                       </body>
                                                                       </html>                        
                                                                       '''

            msg.attach(MIMEText(html, 'html'))

            with open(settings.MEDIA_ROOT + "\\chart.png", 'rb') as f:
                img_data = f.read()
            img = MIMEImage(img_data)
            img.add_header('Content-ID', '<image1>')
            msg.attach(img)

            with open(settings.MEDIA_ROOT + "\\chart2.png", 'rb') as f:
                img_data = f.read()
            img = MIMEImage(img_data)
            img.add_header('Content-ID', '<image2>')
            msg.attach(img)

            with open(settings.MEDIA_ROOT + "\\RedBot_Q1_Apr_Jun_2023.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q1_Apr_Jun_2023.xlsx"))

            with open(settings.MEDIA_ROOT + "\\RedBot_Q2_July_Sept_2023.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q2_July_Sept_2023.xlsx"))

            with open(settings.MEDIA_ROOT + "\\RedBot_Q3_Oct_Dec_2023.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q3_Oct_Dec_2023.xlsx"))

            with open(settings.MEDIA_ROOT + "\\RedBot_Q4_Jan_Mar_2024.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q4_Jan_Mar_2024.xlsx"))

                # RedBot_Q3_Oct_Dec_2023

            smtp_server = 'smtp.office365.com'
            smtp_port = 587
            smtp_username = 'automation@redingtongroup.com'
            smtp_password = '!Redb0t23#'
            smtp_conn = smtplib.SMTP(smtp_server, smtp_port)
            smtp_conn.starttls()
            smtp_conn.login(smtp_username, smtp_password)
            smtp_conn.sendmail(msg['From'], recipients, msg.as_string())
            smtp_conn.quit()
            done = 1
        except Exception as e:
            logger.info("Mail not sent to team" + str(e))
            pass

    return "Notificationdone"











