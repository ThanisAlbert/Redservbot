import datetime
import shutil
import smtplib
from datetime import date
import os
import zipfile
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import json
from urllib.parse import quote
import requests



import xlwt
from django.db import connection
from django.db.models import Q, IntegerField, Max
from django.db.models.functions import Cast
from openpyxl import Workbook
from tablib import Dataset
import sqlite3

from .Bot.History import BotHistory
from .Bot.sqlite import sqlite
from .models import Bot, Process, Subprocess, Workstatus, Botstatus, Kaizenstatus, Kaizenawardedyear, Developermail, \
    Mailreport_to, Mailreport_cc, dbMailrecipient, BotHist, CastAsInteger, TicketTrackingTable
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.utils.dateparse import parse_date

import pyodbc
import logging
logger = logging.getLogger('django')

import pandas as pd

from .tasks import mail_admin


# Create your views here.

def mail_admin(self):
    from datetime import datetime
    import time
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    import smtplib
    from botcreate.models import Bot
    import logging
    from selenium.webdriver.chrome.options import Options
    from selenium import webdriver
    from email.mime.application import MIMEApplication
    from email.mime.image import MIMEImage

    logger = logging.getLogger(__name__)

    message = "Please find the attached Daily Automation Report (01-July-2023 to 30-September-2023 - Q2)"

    logger.info("Testing")

    # =========================================================ChartPreparation======================================

    tickettrackingtable = TicketTrackingTable.objects.all().values()

    #for ticketTracking in tickettrackingtable:
        #print(ticketTracking)

    '''
    options = Options()
    options.headless = False  # Run Chrome in headless mode

    url = "http://172.26.1.19:85/bot/mailchart.html"
    # url = "http://127.0.0.1:8000/bot/mailchart.html"

    driver = webdriver.Chrome(executable_path="E:\\chromedriver.exe", options=options)
    # driver = webdriver.Chrome(executable_path="D:\\chromedriver.exe", options=options)

    driver.get(url)

    # Wait for the chart to be generated (replace "chart-element-id" with the actual ID of the chart element)
    chart_element = driver.find_element_by_id("mailchartcontainer")

    time.sleep(3)

    chart_image = chart_element.screenshot_as_png

    with open(settings.MEDIA_ROOT + "\\chart.png", "wb") as f:
        f.write(chart_image)

        # ===================================Chart2==========================================

        url = "http://172.26.1.19:85/bot/mailchart2.html"
        # url = "http://127.0.0.1:8000/bot/mailchart2.html"

        driver = webdriver.Chrome(executable_path="E:\\chromedriver.exe", options=options)
        # driver = webdriver.Chrome(executable_path="D:\\chromedriver.exe", options=options)

        driver.get(url)

        # Wait for the chart to be generated (replace "chart-element-id" with the actual ID of the chart element)
        chart_element = driver.find_element_by_id("mailchartcontainer")

        time.sleep(3)

        chart_image = chart_element.screenshot_as_png

        with open(settings.MEDIA_ROOT + "\\chart2.png", "wb") as f:
            f.write(chart_image)
    
    '''

    # ==============================================================Q4 Starts======================================================
    from_date_q4 = "2024-01-01"
    to_date_q4 = "2024-03-31"

    #from_date_q4 = datetime.strptime(from_date_q4, "%Y-%m-%d")
    #to_date_q4 = datetime.strptime(to_date_q4, "%Y-%m-%d")


    New_Bots_q4 = TicketTrackingTable.objects.filter(
        Q(startdate__range=(from_date_q4, to_date_q4)) | Q(enddate__range=(from_date_q4, to_date_q4)),
        status__in=["Gone live"],projecttype='Automation').order_by('developermailid')

    Enh_Bots_q4 = TicketTrackingTable.objects.filter(
        Q(enhancestartdate__range=(from_date_q4, to_date_q4)) | Q(enhanceenddate__range=(from_date_q4, to_date_q4)),
        status__in=["Gone live"],projecttype='Automation').order_by('developermailid')

    rest_Bots_q4 = TicketTrackingTable.objects.filter(
        Q(creationdate__range=(from_date_q4, to_date_q4)) | Q(startdate__range=(from_date_q4, to_date_q4)) | Q(
            enddate__range=(from_date_q4, to_date_q4)) | Q(enhancestartdate__range=(from_date_q4, to_date_q4)) | Q(
            enhanceenddate__range=(from_date_q4, to_date_q4)),
        status__in=["Yet to Start", "WIP", "UAT"],projecttype='Automation').order_by('developermailid')

    excel_reports_q4 = TicketTrackingTable.objects.filter(
        Q(creationdate__range=(from_date_q4, to_date_q4)) | Q(startdate__range=(from_date_q4, to_date_q4)) | Q(
            enddate__range=(from_date_q4, to_date_q4)) | Q(enhancestartdate__range=(from_date_q4, to_date_q4)) | Q(
            enhanceenddate__range=(from_date_q4, to_date_q4)),projecttype='Automation').order_by('developermailid')

    table_row_data_q4 = ""
    enhance_table_row_data_q4 = ""
    rest_row_data_q4 = ""

    if New_Bots_q4.exists():

        for obj in New_Bots_q4:

            Botno = obj.projectno
            Businessunit = obj.businessunit
            Botname = obj.projectname
            Botdesc = obj.projectdesc
            Process = obj.process
            Subprocess = obj.subprocess
            Spocname = obj.spocname

            Startdate = ""
            if obj.startdate:
                Startdate = datetime.strptime(str(obj.startdate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.enddate:
                Enddate = datetime.strptime(str(obj.enddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            Botstatus = obj.status
            Developermail = obj.developermailid

            manualtimetaken = ""
            if obj.manualtime and str(obj.manualtime) != "None":
                manualtimetaken = obj.manualtime

            bottimetaken = ""
            if obj.automationtime and str(obj.automationtime) != "None":
                bottimetaken = obj.automationtime

            totaltimesave = ""
            mandaysaved = ""
            if obj.totaltime and str(obj.totaltime) != "None":
                totaltimesave = obj.totaltime

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data_q4 = table_row_data_q4 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data_q4 = "No_data"

    if Enh_Bots_q4.exists():

        for obj in Enh_Bots_q4:

            Botno = obj.projectno
            Businessunit = obj.businessunit
            Botname = obj.projectname
            Botdesc = obj.projectdesc
            Process = obj.process
            Subprocess = obj.subprocess
            Spocname = obj.spocname

            Startdate = ""
            if obj.startdate:
                Startdate = datetime.strptime(str(obj.startdate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.enddate:
                Enddate = datetime.strptime(str(obj.enddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d %H:%M:%S%z").strftime(
                    "%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            Botstatus = obj.status
            Developermail = obj.developermailid

            manualtimetaken = ""
            if obj.manualtime and str(obj.manualtime) != "None":
                manualtimetaken = obj.manualtime

            bottimetaken = ""
            if obj.automationtime and str(obj.automationtime) != "None":
                bottimetaken = obj.automationtime

            totaltimesave = ""
            mandaysaved = ""
            if obj.totaltime and str(obj.totaltime) != "None":
                totaltimesave = obj.totaltime

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data_q4 = enhance_table_row_data_q4 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data_q4 = "No_data"

    if rest_Bots_q4.exists():

        for obj in rest_Bots_q4:

            Botno = obj.projectno
            Businessunit = obj.businessunit
            Botname = obj.projectname
            Botdesc = obj.projectdesc
            Process = obj.process
            Subprocess = obj.subprocess
            Spocname = obj.spocname

            Startdate = ""
            if obj.startdate:
                print(obj.startdate)
                Startdate = datetime.strptime(str(obj.startdate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.enddate:
                Enddate = datetime.strptime(str(obj.enddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d %H:%M:%S%z").strftime(
                    "%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")

            Botstatus = obj.status
            Developermail = obj.developermailid

            manualtimetaken = ""
            if obj.manualtime and str(obj.manualtime) != "None":
                manualtimetaken = obj.manualtime

            bottimetaken = ""
            if obj.automationtime and str(obj.automationtime) != "None":
                bottimetaken = obj.automationtime

            totaltimesave = ""
            mandaysaved = ""
            if obj.totaltime and str(obj.totaltime) != "None":
                totaltimesave = obj.totaltime

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data_q4 = rest_row_data_q4 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data_q4 = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
                'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports_q4:

        if str(obj.creationdate) != "None":
            creationdate = datetime.strptime(str(obj.creationdate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.startdate) != "None":
            startdate = datetime.strptime(str(obj.startdate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.enddate) != "None":
            enddate = datetime.strptime(str(obj.enddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d %H:%M:%S%z").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d %H:%M:%S%z").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.totaltime and str(obj.totaltime) != "None":
            totaltimesave = obj.totaltime

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.projectno, obj.projectname, obj.process, obj.subprocess, obj.spocname, obj.requestormailid, obj.teamleadmailid,
               obj.developermailid, obj.technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.manualtime, obj.automationtime, obj.totaltime, obj.totalday,
               mandaysaved,
               obj.kaizenstatus, obj.status, obj.projectdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q4_Jan_Mar_2024.xlsx")

    # ==============================================================q4 Ends========================================================

    # =============================================================Q3 Starts======================================================

    from_date_q3 = "2023-10-01"
    to_date_q3 = "2023-12-31"

    New_Bots_q3 = Bot.objects.filter(
        Q(Startdate__range=(from_date_q3, to_date_q3)) | Q(Enddate__range=(from_date_q3, to_date_q3)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots_q3 = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date_q3, to_date_q3)) | Q(enhanceenddate__range=(from_date_q3, to_date_q3)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots_q3 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q3, to_date_q3)) | Q(Startdate__range=(from_date_q3, to_date_q3)) | Q(
            Enddate__range=(from_date_q3, to_date_q3)) | Q(enhancestartdate__range=(from_date_q3, to_date_q3)) | Q(
            enhanceenddate__range=(from_date_q3, to_date_q3)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports_q3 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q3, to_date_q3)) | Q(Startdate__range=(from_date_q3, to_date_q3)) | Q(
            Enddate__range=(from_date_q3, to_date_q3)) | Q(enhancestartdate__range=(from_date_q3, to_date_q3)) | Q(
            enhanceenddate__range=(from_date_q3, to_date_q3))).order_by('Developermail')

    table_row_data_q3 = ""
    enhance_table_row_data_q3 = ""
    rest_row_data_q3 = ""

    if New_Bots_q3.exists():

        for obj in New_Bots_q3:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data_q3 = table_row_data_q3 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data_q3 = "No_data"

    if Enh_Bots_q3.exists():

        for obj in Enh_Bots_q3:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data_q3 = enhance_table_row_data_q3 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data_q3 = "No_data"

    if rest_Bots_q3.exists():

        for obj in rest_Bots_q3:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data_q3 = rest_row_data_q3 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data_q3 = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports_q3:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q3_Oct_Dec_2023.xlsx")

    # =============================================================Q3 Ends=========================================================

    # ==============================================================Q2 Starts======================================================

    from_date = "2023-07-01"
    to_date = "2023-09-30"

    New_Bots = Bot.objects.filter(Q(Startdate__range=(from_date, to_date)) | Q(Enddate__range=(from_date, to_date)),
                                  Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date, to_date)) | Q(enhanceenddate__range=(from_date, to_date)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots = Bot.objects.filter(
        Q(Creationdate__range=(from_date, to_date)) | Q(Startdate__range=(from_date, to_date)) | Q(
            Enddate__range=(from_date, to_date)) | Q(enhancestartdate__range=(from_date, to_date)) | Q(
            enhanceenddate__range=(from_date, to_date)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports = Bot.objects.filter(
        Q(Creationdate__range=(from_date, to_date)) | Q(Startdate__range=(from_date, to_date)) | Q(
            Enddate__range=(from_date, to_date)) | Q(enhancestartdate__range=(from_date, to_date)) | Q(
            enhanceenddate__range=(from_date, to_date))).order_by('Developermail')

    table_row_data = ""
    enhance_table_row_data = ""
    rest_row_data = ""

    if New_Bots.exists():

        for obj in New_Bots:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data = table_row_data + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data = "No_data"

    if Enh_Bots.exists():

        for obj in Enh_Bots:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data = enhance_table_row_data + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data = "No_data"

    if rest_Bots.exists():

        for obj in rest_Bots:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data = rest_row_data + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q2_July_Sept_2023.xlsx")

    # ==========================================================q2 ends=================================================

    # ===========================================================q1 starts===============================================

    from_date_q1 = "2023-04-01"
    to_date_q1 = "2023-06-30"

    New_Bots_q1 = Bot.objects.filter(
        Q(Startdate__range=(from_date_q1, to_date_q1)) | Q(Enddate__range=(from_date_q1, to_date_q1)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    Enh_Bots_q1 = Bot.objects.filter(
        Q(enhancestartdate__range=(from_date_q1, to_date_q1)) | Q(enhanceenddate__range=(from_date_q1, to_date_q1)),
        Botstatus__in=["Completed"]).order_by('Developermail')

    rest_Bots_q1 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q1, to_date_q1)) | Q(Startdate__range=(from_date_q1, to_date_q1)) | Q(
            Enddate__range=(from_date_q1, to_date_q1)) | Q(enhancestartdate__range=(from_date_q1, to_date_q1)) | Q(
            enhanceenddate__range=(from_date_q1, to_date_q1)),
        Botstatus__in=["Bot No.Generated", "Devolopment In Progress", "TM TO BE DONE", "Under User Testing"]).order_by(
        'Developermail')

    excel_reports_q1 = Bot.objects.filter(
        Q(Creationdate__range=(from_date_q1, to_date_q1)) | Q(Startdate__range=(from_date_q1, to_date_q1)) | Q(
            Enddate__range=(from_date_q1, to_date_q1)) | Q(enhancestartdate__range=(from_date_q1, to_date_q1)) | Q(
            enhanceenddate__range=(from_date_q1, to_date_q1))).order_by('Developermail')

    table_row_data_q1 = ""
    enhance_table_row_data_q1 = ""
    rest_row_data_q1 = ""

    if New_Bots_q1.exists():

        for obj in New_Bots_q1:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            table_row_data_q1 = table_row_data_q1 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                totaltimesave) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        table_row_data_q1 = "No_data"

    if Enh_Bots_q1.exists():

        for obj in Enh_Bots_q1:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            enhance_table_row_data_q1 = enhance_table_row_data_q1 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"


    else:
        enhance_table_row_data_q1 = "No_data"

    if rest_Bots_q1.exists():

        for obj in rest_Bots_q1:

            Botno = obj.Botno
            Businessunit = obj.businessunit
            Botname = obj.Botname
            Botdesc = obj.Botdesc
            Process = obj.Process
            Subprocess = obj.Subprocess
            Spocname = obj.Spocname

            Startdate = ""
            if obj.Startdate:
                Startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Enddate = ""
            if obj.Enddate:
                Enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhancestartdate = ""
            if obj.enhancestartdate:
                enhancestartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")

            enhanceenddate = ""
            if obj.enhanceenddate:
                enhanceenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")

            Botstatus = obj.Botstatus
            Developermail = obj.Developermail

            manualtimetaken = ""
            if obj.Manualtimespend and str(obj.Manualtimespend) != "None":
                manualtimetaken = obj.Manualtimespend

            bottimetaken = ""
            if obj.Automationtimespend and str(obj.Automationtimespend) != "None":
                bottimetaken = obj.Automationtimespend

            totaltimesave = ""
            mandaysaved = ""
            if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
                totaltimesave = obj.Totaltimesaved

                if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                    mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                    if mandaysaved <= 8:
                        mandaysaved = str(mandaysaved) + " hrs"
                    else:
                        mandaysaved = str(round(mandaysaved / 8)) + " Business days"
                else:
                    mandaysaved = totaltimesave

            rest_row_data_q1 = rest_row_data_q1 + "<tr><td><a style='color:black' href='http://172.26.1.19:85/bot/editbot.html/" + str(
                Botno) + "'>" + str(Botno) + "</a></td><td>" + str(
                Businessunit) + "</td><td style='white-space:nowrap';>" + str(Botname) + "</td><td>" + str(
                Botdesc) + "</td><td>" + str(
                Spocname) + "</td><td style='color:green'>" + str(
                manualtimetaken) + "</td><td style='color:green'>" + str(
                bottimetaken) + "</td><td style='color:green'>" + str(mandaysaved) + "</td><td>" + str(
                Developermail) + "</td><td>" + str(Subprocess) + "</td><td>" + str(
                Startdate) + "</td><td>" + str(
                Enddate) + "</td><td>" + str(enhancestartdate) + "</td><td>" + str(
                enhanceenddate) + "</td><td>" + str(Botstatus) + "</td><td>"

    else:
        rest_row_data_q1 = "No_data"

    wb = Workbook()
    ws = wb.active

    headers = ['Botno', 'Botname', 'Process', 'Subprocess', 'Spocname', 'RequestBy', 'Teamlead', 'DevelopmentBy',
               'Technologyused', 'Creationdate', 'Startdate', 'Enddate', 'Enhancementstartdate', 'Enhancementenddate',
               'Manualtimespend', 'Automationtimespend', 'Totaltimesaved', 'Totaldaysavedinmins',
               'Total-man-day-savingsquarterly', 'Kaizenawardstatus',
               'Kaizenawardedyear', 'Botstatus', 'Botdesc']

    ws.append(headers)

    for obj in excel_reports_q1:
        if str(obj.Creationdate) != "None":
            creationdate = datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            creationdate = ""

        if str(obj.Startdate) != "None":
            startdate = datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            startdate = ""

        if str(obj.Enddate) != "None":
            enddate = datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enddate = ""

        if str(obj.enhancestartdate) != "None":
            enhancementstartdate = datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime(
                "%d-%m-%Y")
        else:
            enhancementstartdate = ""

        if str(obj.enhanceenddate) != "None":
            enhancementenddate = datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            enhancementenddate = ""

        mandaysaved = ""
        if obj.Totaltimesaved and str(obj.Totaltimesaved) != "None":
            totaltimesave = obj.Totaltimesaved

            if "mins" in str(totaltimesave) and "days" not in str(totaltimesave):
                mandaysaved = round(int(str(totaltimesave).replace("mins", "")) / 60)
                if mandaysaved <= 8:
                    mandaysaved = str(mandaysaved) + " hrs"
                else:
                    mandaysaved = str(round(mandaysaved / 8)) + " Business days"
            else:
                mandaysaved = totaltimesave

        row = [obj.Botno, obj.Botname, obj.Process, obj.Subprocess, obj.Spocname, obj.Requestormail, obj.Teamleadmail,
               obj.Developermail, obj.Technologyused, creationdate, startdate, enddate, enhancementstartdate,
               enhancementenddate, obj.Manualtimespend, obj.Automationtimespend, obj.Totaltimesaved, obj.Totaldaysaved,
               mandaysaved,
               obj.Kaizenawardstatus, obj.Kaizenawardyear, obj.Botstatus, obj.Botdesc]

        ws.append(row)

    wb.save(settings.MEDIA_ROOT + "\\RedBot_Q1_Apr_Jun_2023.xlsx")

    # ============================================================q1 ends=================================================

    done = 0

    while done == 0:
        try:
            msg = MIMEMultipart()
            msg['From'] = 'botomation@redingtongroup.com'

            to_addr = []
            # to_addr.append("prajnya.sahu@redingtongroup.com")
            # to_addr.append("prasanth.muruga@redingtongroup.com")
            # to_addr.append("santhosh.arjunan@redingtongroup.com")
            to_addr.append("thanis.albert@redingtongroup.com")

            cc_addr = []
            # cc_addr.append("rathina.moorthy@redingtongroup.com")
            # cc_addr.append("raheema.shahul@redingtongroup.com")
            # cc_addr.append("srinivasa.babu@redingtongroup.com")

            bc_addr = []
            bc_addr.append("thanis.a@redingtongroup.com")

            msg['To'] = ', '.join(to_addr)
            msg['Cc'] = ', '.join(cc_addr)
            msg['Bcc'] = ', '.join(bc_addr)

            recipients = to_addr + bc_addr + cc_addr

            msg['Subject'] = 'Daily Automation Report (01-Jan-2024 to 31-March-2024 - Q4)'

            html = '''
                                                                       <html>
                                                                       <style>
                                                                         table,tr,td {
                                                                         border: 1px solid black;
                                                                         border-collapse:collapse;                          
                                                                         }
                                                                         span{
                                                                         color:red;
                                                                         }
                                                                       </style>
                                                                       <body>                                                                   

                                                                       Dear Team,
                                                                       <br><br>

                                                                       <br>  

                                                                       <img src="cid:image2">

                                                                       <br><br><br><br><br>

                                                                       <img src="cid:image1">

                                                                       <br><br><br><br>
                                                                       Please find the attached Daily Automation Report (01-January-2024 to 31-March-2024 - Q4)
                                                                       <br> 


                                                                       <br><br>*************************************************************************<b style="color:green;">(01-January-2024 to 31-March-2024 - Q4)</b>******************************************************<br><br><br>

                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data_q4 + '''                                             
                                                                       </table> 

                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data_q4 + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data_q4 + '''                                             
                                                                       </table>


                                                                       <br><br>*************************************************************************<b style="color:green;">(01-October-2023 to 31-December-2023 - Q3)</b>******************************************************<br><br><br>

                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data_q3 + '''                                             
                                                                       </table> 

                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data_q3 + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data_q3 + '''                                             
                                                                       </table>



                                                                        <br><br><br><br><br>*************************************************************************<b style="color:green;">(01-July-2023 to 30-September-2023 - Q2)</b>********************************************************<br><br><br><br>


                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data + '''                                             
                                                                       </table>                                                                   


                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data + '''                                             
                                                                       </table>

                                                                       <br><br>
                                                                       <br><br>*************************************************************************<b style="color:green;">(01-April-2023 to 31-June-2023 - Q1)</b>******************************************************<br><br><br>

                                                                       <u><i>Enhancement Bots (Completed)</i></u><br><br>

                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + enhance_table_row_data_q1 + '''                                             
                                                                       </table> 

                                                                       <br><br><u><i>New Bots (Completed)</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                           <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                            
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus</b></td>  


                                                                         </tr>
                                                                         ''' + table_row_data_q1 + '''                                             
                                                                       </table>


                                                                       <br><br><u><i>Development In Progress Bots Detail</i></u><br><br>
                                                                       <table>
                                                                         <tr>
                                                                           <td style="background-color:#F0F0F0;"><b>Botno</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>BusinessUnit</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                    
                                                                            <td style="background-color:#F0F0F0;"><b>Description&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Spocname&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>

                                                                           <td style="background-color:#F0F0F0;"><b>ManualTime</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>BotTimeTaken</b></td>                                                                          
                                                                           <td style="background-color:#F0F0F0;"><b>Total-man-day-savingsquarterly</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Developer</b></td> 
                                                                           <td style="background-color:#F0F0F0;"><b>Subprocess&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>                                                                      
                                                                           <td style="background-color:#F0F0F0;"><b>Startdate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enddate&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhancestartdate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Enhanceenddate </b></td>
                                                                           <td style="background-color:#F0F0F0;"><b>Botstatus&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</b></td>  


                                                                         </tr>
                                                                         ''' + rest_row_data_q1 + '''                                             
                                                                       </table>




                                                                       <br>                                                           
                                                               Bot Application Link: <a href ="http://172.26.1.19:85/bot/">http://172.26.1.19:85/bot/</a>                                               
                                                               <br><br>This is an automatically generated email. <br><br>
                                                                       Thanks & Regards,<br>
                                                                       Automation Team - RGS
                                                                       </body>
                                                                       </html>                        
                                                                       '''

            msg.attach(MIMEText(html, 'html'))

            with open(settings.MEDIA_ROOT + "\\chart.png", 'rb') as f:
                img_data = f.read()
            img = MIMEImage(img_data)
            img.add_header('Content-ID', '<image1>')
            msg.attach(img)

            with open(settings.MEDIA_ROOT + "\\chart2.png", 'rb') as f:
                img_data = f.read()
            img = MIMEImage(img_data)
            img.add_header('Content-ID', '<image2>')
            msg.attach(img)

            with open(settings.MEDIA_ROOT + "\\RedBot_Q1_Apr_Jun_2023.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q1_Apr_Jun_2023.xlsx"))

            with open(settings.MEDIA_ROOT + "\\RedBot_Q2_July_Sept_2023.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q2_July_Sept_2023.xlsx"))

            with open(settings.MEDIA_ROOT + "\\RedBot_Q3_Oct_Dec_2023.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q3_Oct_Dec_2023.xlsx"))

            with open(settings.MEDIA_ROOT + "\\RedBot_Q4_Jan_Mar_2024.xlsx", 'rb') as file:
                msg.attach(MIMEApplication(file.read(), Name="RedBot_Q4_Jan_Mar_2024.xlsx"))

                # RedBot_Q3_Oct_Dec_2023

            smtp_server = 'smtp.office365.com'
            smtp_port = 587
            smtp_username = 'automation@redingtongroup.com'
            smtp_password = '!Redb0t23#'
            smtp_conn = smtplib.SMTP(smtp_server, smtp_port)
            smtp_conn.starttls()
            smtp_conn.login(smtp_username, smtp_password)
            smtp_conn.sendmail(msg['From'], recipients, msg.as_string())
            smtp_conn.quit()
            done = 1
        except Exception as e:
            logger.info("Mail not sent to team" + str(e))
            pass

    return "Notificationdone"


#===========================BotMonitor===========================

def botMonitor(request):

    server = '172.26.1.21'
    database = 'botomation'
    username = 'botuser'
    password = 'botuser@27!'
    conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'

    # Establish the database connection
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # Execute a SQL query to read records from a table
    cursor.execute('SELECT * FROM bot_tracker_table')

    botsqlresult = [];

    # Fetch the results and print them
    for row in cursor.fetchall():
        botsqlresult.append(row)

    # Close the cursor and connection
    cursor.close()
    conn.close()
    context = {'bots': botsqlresult}

    return render(request, 'bot/bot_monitor.html', context)



#=============================Index================================================

def login(request):
    #external_url = "http://172.24.3.13:88/"  # Replace with the actual URL
    #return redirect(external_url)
    try:
        logger.info(str("last logged in at ") + str(datetime.datetime.now()))
    except Exception as e:
        print (e)

    return render(request, 'bot/login.html')

def botview(request):
    return render(request, 'bot/botview.html')

def loginProcess(request):
    username = request.POST["username"]
    password = request.POST["pass"]

    if str(username).upper() == "BOT" and str(password).upper() == "BOT":
        return redirect('bot:botviewquery')
        #request.session["loginvar"] = "allow"
    else:
        return redirect('bot:login')

def logout(request):
    #request.session["loginvar"] = ""
    return redirect('bot:login')

def index(request):

    loginvar = "allow"

    if loginvar == "allow":

        try:

            db = sqlite3.connect('db.sqlite3')

            # Create a cursor object to execute SQL queries
            cursor = db.cursor()

            # Execute a query to get all table names
            cursor.execute("SELECT * from botcreate_process;")
            # Fetch all rows from the result set
            process = cursor.fetchall()

            cursor.execute("SELECT * from botcreate_subprocess;")
            subprocess = cursor.fetchall()

            cursor.execute("SELECT * from botcreate_subprocess;")
            subprocess = cursor.fetchall()

            cursor.execute("SELECT * from botcreate_botstatus;")
            botstatus = cursor.fetchall()

            cursor.execute("SELECT * from botcreate_kaizenstatus;")
            kaizenstatus = cursor.fetchall()

            cursor.execute("SELECT * from botcreate_developermail;")
            developermail = cursor.fetchall()

            cursor.execute("SELECT * from botcreate_workstatus;")
            livestatus = cursor.fetchall()

            newbotno=""

            '''
            #db = sqlite()
            #process = db.process()
            subprocess = db.subprocess()
            botstatus = db.botstatus()
            kaizenstatus = db.kaizenstatus()
            developermail = db.developermail()
            newbotno = db.newbotno()
            livestatus = db.livestatus()
            '''

            cursor.close()
            db.close()


        except Exception as e:
            print(e)


        try:
            Bots = TicketTrackingTable.objects.filter(Botstatus=request.POST["dropdownselect"]).order_by('projectno')
        except:
            Bots = TicketTrackingTable.objects.filter(
                Q(status='WIP') | Q(status='UAT') | Q(status='Yet to Start'), projecttype='Automation' ).order_by('-projectno').values()

        for bot in Bots:

            if str(bot["startdate"]) != "None":
                bot["startdate"] = str(bot["startdate"]).split(" ")[0]
                bot["startdate"] = datetime.datetime.strptime(bot["startdate"], "%Y-%m-%d")
                bot["startdate"] = bot["startdate"].strftime("%b %d, %Y")

            if  str(bot["enddate"]) !="None":
                bot["enddate"] = str(bot["enddate"]).split(" ")[0]
                bot["enddate"] = datetime.datetime.strptime(bot["enddate"], "%Y-%m-%d")
                bot["enddate"] = bot["enddate"].strftime("%b %d, %Y")


        context = {'newbotno': newbotno,  'kaizenstatus':kaizenstatus, 'livestatus':livestatus ,'botstatus':botstatus ,'process': process,'subprocess': subprocess,'developermail': developermail,'Bots': Bots}

        return render(request, 'bot/index.html', context)

    else:
        return render(request, 'bot/login.html')


def editbot(request,param):

    loginvar = "allow"

    if loginvar == "allow":

        Botno = param

        Bots = TicketTrackingTable.objects.filter(
            Q(status='WIP') | Q(status='UAT') | Q(status='Yet to Start'), projecttype='Automation').order_by('-projectno').values()

        Botobj = TicketTrackingTable.objects.filter(projectno=Botno).values()

        if Botobj.exists():
            pass
        else:
            return redirect('bot:index')

        db = sqlite()
        process = db.process()
        subprocess = db.subprocess()
        botstatus = db.botstatus()
        kaizenstatus = db.kaizenstatus()
        developermail = db.developermail()
        newbotno = db.newbotno()
        livestatus = db.livestatus()

        for bot in Bots:

            if str(bot["startdate"]) != "None":
                bot["startdate"] = str(bot["startdate"]).split(" ")[0]
                bot["startdate"] = datetime.datetime.strptime(bot["startdate"], "%Y-%m-%d")
                bot["startdate"] = bot["startdate"].strftime("%b %d, %Y")

            if str(bot["enddate"]) != "None":
                bot["enddate"] = str(bot["enddate"]).split(" ")[0]
                bot["enddate"] = datetime.datetime.strptime(bot["enddate"], "%Y-%m-%d")
                bot["enddate"] = bot["enddate"].strftime("%b %d, %Y")

        Context = {'newbotno': newbotno, 'Botdetails': Botobj, 'developermail': developermail, 'Bots': Bots, 'process': process,
                   'subprocess': subprocess, 'botstatus': botstatus, 'kaizenstatus': kaizenstatus,
                   'kaizenawardedyear': livestatus }

        return render(request, 'bot/index_search.html', Context)
    else:
        return render(request, 'bot/login.html')

def createbot(request):

    loginvar = "allow"

    if loginvar == "allow":

        if 'submit' in request.POST:
            Botno = request.POST['botno']
            Botname = request.POST["botname"]
            Process_var = request.POST["process"]
            Subprocess_var = request.POST["subprocess"]
            Spocname = request.POST["spocname"]
            Requestormail = request.POST["requestormail"]
            Teamleadmail = request.POST["teamleadmail"]
            Managermail = request.POST["managermail"]
            Developermail_var = request.POST["developermail"]
            Technologyused = request.POST["technology"]

            try:
                Creationdate = datetime.datetime.strptime(request.POST["creationdate"], '%d/%m/%Y')
            except:
                Creationdate = ""

            try:
                Startdate = datetime.datetime.strptime(request.POST["startdate"], '%d/%m/%Y')
            except:
                Startdate = ""

            try:
                Enddate = datetime.datetime.strptime(request.POST["enddate"], '%d/%m/%Y')
            except:
                Enddate = ""

            try:
                enhancestartdate = datetime.datetime.strptime(request.POST["enhancestart"], '%d/%m/%Y')
            except:
                enhancestartdate = ""

            try:
                enhanceenddate = datetime.datetime.strptime(request.POST["enhanceend"], '%d/%m/%Y')
            except:
                enhanceenddate = ""

            Botstatus_var = request.POST["botstatus"]
            Manualtime = request.POST["manualtime"]
            Automationtime = request.POST["automationtime"]
            Totaltime = request.POST["totaltime"]
            Totalday = request.POST["totalday"]
            Kaizenstatus_var = request.POST["kaizenstatus"]
            Kaizenyear = request.POST["kaizenyear"]
            Botdesc = request.POST["botdesc"]
            Mailrecipient = request.POST["mailrecipient"]
            businessUnit = request.POST["businessunit"]
            priority = request.POST["priority"]
            categorization = request.POST["categorization"]
            remarks_var = request.POST["remarks"]

            bot = Bot()
            bot.Botno = Botno
            bot.Botname = Botname
            bot.Process = Process_var
            bot.Subprocess = Subprocess_var
            bot.Spocname = Spocname
            bot.Requestormail = Requestormail
            bot.Teamleadmail = Teamleadmail
            bot.Managermail = Managermail
            bot.Developermail = Developermail_var
            bot.Technologyused = Technologyused
            try:
                bot.Creationdate = Creationdate.strftime("%Y-%m-%d")
            except:
                bot.Creationdate = None

            try:
                bot.Startdate = Startdate.strftime("%Y-%m-%d")
            except:
                bot.Startdate = None

            try:
                bot.Enddate = Enddate.strftime("%Y-%m-%d")
            except:
                bot.Enddate = None

            try:
                bot.enhancestartdate = enhancestartdate.strftime("%Y-%m-%d")
            except:
                bot.enhancestartdate =None

            try:
                bot.enhanceenddate = enhanceenddate.strftime("%Y-%m-%d")
            except:
                bot.enhanceenddate = None

            bot.Botstatus = Botstatus_var
            bot.Manualtimespend = Manualtime
            bot.Automationtimespend = Automationtime
            bot.Totaltimesaved = Totaltime
            bot.Totaldaysaved = Totalday
            bot.Kaizenawardstatus = Kaizenstatus_var
            #bot.Kaizenawardyear = Kaizenyear
            bot.livestatus=Kaizenyear
            bot.priority = priority
            bot.categorization = categorization
            bot.Botdesc = Botdesc
            bot.Mailrecipient = Mailrecipient
            bot.businessunit=businessUnit
            bot.remarks=remarks_var


            Botobj = Bot.objects.filter(Botno=Botno).values()

            if Botobj.exists():
                request.session['botcreateerror'] = "Bot already exist"
                return redirect('bot:index')
            else:
                bot.save()
                try:
                    #myfile = request.FILES['myfile']
                    myfiles = request.FILES.getlist('myfile')

                    for myfile in myfiles:
                        fs = FileSystemStorage()
                        folder_name = Botno
                        folder_path = os.path.join("E:\\Botomation\\data\\storage1\\", folder_name)
                        #folder_path = os.path.join("C:\\Automation\\Python\\Redservbot\\Redservbot\\media", folder_name)
                        if not os.path.exists(folder_path):
                            os.makedirs(folder_path)
                        file = fs.save(os.path.join(folder_name, myfile.name), myfile)
                except:
                    pass

                msg = MIMEMultipart('related')
                name = str(str(Requestormail).split("@")[0]).replace(".","")
                MESSAGE_BODY = """\
                    <html>
                      <head>
                        <style>
                        table, th, td {
      border: 1px solid black;
      border-collapse: collapse;  
    }
    th{
    background-color:#C7CBC7;
    text-align: left;
    }
                        </style>
                      </head>
                        <body>
                        
                          Dear Sender <br><br>
                          Greetings for the day!!!<br><br>
                          We have generated Redbot number for the below activity for future reference<br><br>
                          <table>
                            <tr>
                               <th>Bot No.</th>
                               <td>"""+ Botno +"""</td>
                            </tr> 
                            <tr>
                               <th>Process</th>
                               <td>"""+ Process_var+"""</td>
                            </tr>
                            <tr>
                               <th>Sub Process</th>
                               <td>"""+Subprocess_var+"""</td>
                            </tr> 
                            <tr>
                               <th>Bot Name</th>
                               <td>"""+Botname+"""</td>
                            </tr> 
                            <tr>
                               <th>Business Unit</th>
                               <td>""" + str(businessUnit) + """</td>
                            </tr> 
                            <tr>
                               <th>Contact Person</th>
                               <td>"""+ Spocname + """</td>
                            </tr> 
                            <tr>
                               <th>Developer Name</th>
                               <td>"""+ str(str(Developermail_var).split("@")[0]).replace("."," ") + """</td>
                            </tr>  
                            <tr>
                               <th>Manual Time Spend</th>
                               <td>Yet to be calculated</td>
                            </tr> 
                            <tr>
                               <th>Automation Time Spend</th>
                               <td>Yet to be calculated</td>
                            </tr>                    
                          </table>
                          <br>
                          Thanks and Regards,<br>Team Automation
                        </body>
                    </html>
                    """

                to_addr = []
                to_addr.append(Requestormail)
                cc_addr=[]
                cc_addr.append(Teamleadmail)
                cc_addr.append(Managermail)
                cc_addr.append(Developermail_var)
                Mailrecipientres_query = dbMailrecipient.objects.all().values()

                if ";" in str(Mailrecipient):
                    Mailrecipient_list = str(Mailrecipient).split(";")
                    for recipient in Mailrecipient_list:
                        cc_addr.append(recipient)
                else:
                    cc_addr.append(Mailrecipient)

                for res in Mailrecipientres_query:
                    if ";" in str(res["dbmailrecipient"]):
                        Mailrecipient_list = str(res["dbmailrecipient"]).split(";")
                        for recipient in Mailrecipient_list:
                            cc_addr.append(recipient)
                    else:
                        cc_addr.append(res["dbmailrecipient"])


                body_part = MIMEText(MESSAGE_BODY, 'html')

                msg['Subject'] = Botstatus_var + "-" + Botno + "-" + Subprocess_var + "-" +Botname
                #msg['Subject'] = "Bot Status for " + date.today().strftime('%d/%m/%Y')
                msg['From'] = "automation@redingtongroup.com"
                msg['To'] = ', '.join(to_addr)
                msg['Cc'] = ', '.join(cc_addr)
                recipients=to_addr+cc_addr
                msg.attach(body_part)

                server = smtplib.SMTP("smtp.office365.com", 587)
                server.starttls()
                try:
                    server.login("automation@redingtongroup.com", "!Redb0t23#")
                    server.sendmail(msg['From'], recipients, msg.as_string())
                    server.quit()
                except Exception as e:
                    return HttpResponse(e)

                return redirect('bot:index')

        if 'search' in request.POST:

            Botno = request.POST['botno']

            Bots = TicketTrackingTable.objects.filter(
                Q(status='WIP') | Q(status='UAT') | Q(status='Yet to Start'), projecttype='Automation' ).order_by('-projectno').values()

            Botobj = TicketTrackingTable.objects.filter(projectno=Botno).values()

            if Botobj.exists():
                pass
            else:
                return redirect('bot:index')

            db = sqlite()
            process = db.process()
            subprocess = db.subprocess()
            botstatus = db.botstatus()
            kaizenstatus = db.kaizenstatus()
            developermail = db.developermail()
            newbotno = db.newbotno()
            livestatus = db.livestatus()

            for bot in Bots:

                if str(bot["startdate"]) != "None":
                    bot["startdate"] = str(bot["startdate"]).split(" ")[0]
                    bot["startdate"] = datetime.datetime.strptime(bot["startdate"], "%Y-%m-%d")
                    bot["startdate"] = bot["startdate"].strftime("%b %d, %Y")

                if str(bot["enddate"]) != "None":
                    bot["enddate"] = str(bot["enddate"]).split(" ")[0]
                    bot["enddate"] = datetime.datetime.strptime(bot["enddate"], "%Y-%m-%d")
                    bot["enddate"] = bot["enddate"].strftime("%b %d, %Y")

            Context = {'Botdetails':Botobj,'developermail':developermail,'Bots': Bots,'process':process,'subprocess':subprocess,'botstatus':botstatus,'kaizenstatus':kaizenstatus,'kaizenawardedyear':livestatus}

            return render(request,'bot/index_search.html',Context)

        if 'update' in request.POST:

            botno=request.POST['botno']

            try:
                botobj = TicketTrackingTable.objects.get(projectno=botno)
            except:
                return redirect('bot:index')

            Botname = request.POST["botname"]
            Process_var = request.POST["process"]
            Subprocess_var = request.POST["subprocess"]
            Spocname = request.POST["spocname"]
            Requestormail = request.POST["requestormail"]
            Teamleadmail = request.POST["teamleadmail"]
            Managermail = request.POST["managermail"]
            Developermail_var = request.POST["developermail"]
            Technologyused = request.POST["technology"]
            businessUnit = request.POST["businessunit"]
            priority = request.POST["priority"]
            categorization = request.POST["categorization"]
            remarks_var = request.POST["remarks"]


            if Botname=="" and Requestormail=="":
                return redirect('bot:index')

            try:
                Creationdate = datetime.datetime.strptime(request.POST["creationdate"], '%d/%m/%Y')
            except Exception as e:
                Creationdate = ""

            try:
                Startdate = datetime.datetime.strptime(request.POST["startdate"], '%d/%m/%Y')
            except:
                Startdate = ""

            try:
                Enddate = datetime.datetime.strptime(request.POST["enddate"], '%d/%m/%Y')
            except:
                Enddate = ""

            try:
                enhancestartdate = datetime.datetime.strptime(request.POST["enhancestart"], '%d/%m/%Y')
            except:
                enhancestartdate = ""

            try:
                enhanceenddate = datetime.datetime.strptime(request.POST["enhanceend"], '%d/%m/%Y')
            except:
                enhanceenddate = ""


            Botstatus_var = request.POST["botstatus"]
            Manualtime = request.POST["manualtime"]
            Automationtime = request.POST["automationtime"]
            Totaltime = request.POST["totaltime"]
            Totalday = request.POST["totalday"]
            Kaizenstatus_var = request.POST["kaizenstatus"]
            Kaizenyear = request.POST["kaizenyear"]
            Botdesc = request.POST["botdesc"]
            Mailrecipient = request.POST["mailrecipient"]
            #Mailnotes = request.POST["mailnotes"]
            Mailsend = request.POST.getlist('mailsend')

            botobj.projectno = request.POST['botno']
            botobj.projectname = Botname
            botobj.process = Process_var
            botobj.subprocess = Subprocess_var
            botobj.spocname = Spocname
            botobj.requestormailid = Requestormail
            botobj.teamleadmailid = Teamleadmail
            botobj.managermailid = Managermail
            botobj.developermailid = Developermail_var
            botobj.technologyused = Technologyused
            try:
                botobj.creationdate = Creationdate.strftime("%Y-%m-%d")
            except:
                botobj.creationdate = None

            try:
                botobj.startdate = Startdate.strftime("%Y-%m-%d")
            except:
                botobj.startdate = None

            try:
                botobj.enddate = Enddate.strftime("%Y-%m-%d")
            except:
                botobj.enddate = None

            try:
                botobj.enhancestartdate = enhancestartdate.strftime("%Y-%m-%d")
            except:
                botobj.enhancestartdate =None

            try:
                botobj.enhanceenddate = enhanceenddate.strftime("%Y-%m-%d")
            except:
                botobj.enhanceenddate = None

            botobj.status = Botstatus_var
            botobj.manualtime = Manualtime
            botobj.automationtime = Automationtime
            botobj.totaltime = Totaltime
            botobj.totalday = Totalday
            botobj.Kaizenstatus = Kaizenstatus_var
            #botobj.Kaizenawardyear = Kaizenyear
            botobj.livestatus=Kaizenyear
            botobj.priority = priority
            botobj.categorization = categorization
            botobj.projectdesc = Botdesc
            botobj.mailrecipient = Mailrecipient
            botobj.Mailnotes = ""
            botobj.businessunit = businessUnit
            botobj.remarks =remarks_var

            mailcontent = ""

            botdir = "E:\\Botomation\\data\\storage1\\"

            #zf = zipfile.ZipFile(botdir + botno + ".zip", "w")
            #for dirname, subdirs, files in os.walk(botdir):
            #    zf.write(dirname)
            #    for filename in files:
            #        zf.write(os.path.join(dirname, filename))
            #zf.close()

            shutil.make_archive(botdir +botno,'zip',botdir +botno)

            if Botstatus_var == "Bot No.Generated":
                mailcontent = "This is to inform that bot development is under progress."

            if Botstatus_var =="Completed":
                mailcontent = "This is to inform that the bot development has been completed. Please let us know if any changes need to be done."

            if Botstatus_var == "Under User Testing":
                mailcontent = "This is to inform that the bot development has been completed. Please do UAT and let us know if any changes need to be done."

            if Botstatus_var == "Cancelled":
                mailcontent = "We regret to inform that bot development has been cancelled. Please contact us for further clarification"

            if Mailsend:
                msg = MIMEMultipart('related')
                name = str(str(Requestormail).split("@")[0]).replace(".","")
                MESSAGE_BODY = """\
                                <html>
                                  <head>
                                    <style>
                                    table, th, td {
                  border: 1px solid black;
                  border-collapse: collapse;  
                }
                th{
                background-color:#C7CBC7;
                }
                                    </style>
                                  </head>
                                    <body>
    
                                      Dear Sender <br><br>
                                      Greetings for the day!!!<br><br>
                                      
                                      """ + mailcontent + """ <br><br>
                                     
                                      <table>
                                        <tr>
                                           <th>Bot No.</th>
                                           <td>""" + botno + """</td>
                                        </tr> 
                                        <tr>
                                           <th>Process</th>
                                           <td>""" + Process_var + """</td>
                                        </tr>
                                        <tr>
                                           <th>Sub Process</th>
                                           <td>""" + Subprocess_var + """</td>
                                        </tr> 
                                        <tr>
                                           <th>Bot Name</th>
                                           <td>""" + Botname + """</td>
                                        </tr> 
                                        <tr>
                                           <th>Requested By</th>
                                           <td>""" + str(str(Requestormail).split("@")[0]).replace(".","") + """</td>
                                        </tr> 
                                        <tr>
                                           <th>Contact Person</th>
                                           <td>""" + Spocname + """</td>
                                        </tr> 
                                        <tr>
                                           <th>Developer Name</th>
                                           <td>""" + str(str(Developermail_var).split("@")[0]).replace("."," ") + """</td>
                                        </tr>
                                        <tr>
                                        <th>Manual Time Spend</th>
                                           <td>""" + Manualtime + """</td>
                                        </tr>
                                        <tr>
                                        <th>Automation Time Spend</th>
                                           <td>""" + Automationtime + """</td>
                                        </tr>                      
                                      </table>
                                      <br>
                                      Thanks and Regards,<br>Team Automation
                                    </body>
                                </html>
                                """

                to_addr = []
                to_addr.append(Requestormail)
                cc_addr = []
                cc_addr.append(Teamleadmail)
                cc_addr.append(Managermail)
                cc_addr.append(Developermail_var)
                Mailrecipientres_query = dbMailrecipient.objects.all().values()

                if ";" in str(Mailrecipient):
                    Mailrecipient_list = str(Mailrecipient).split(";")
                    for recipient in Mailrecipient_list:
                        cc_addr.append(recipient)
                else:
                    cc_addr.append(Mailrecipient)

                for res in Mailrecipientres_query:
                    if ";" in str(res["dbmailrecipient"]):
                        Mailrecipient_list = str(res["dbmailrecipient"]).split(";")
                        for recipient in Mailrecipient_list:
                            cc_addr.append(recipient)
                    else:
                        cc_addr.append(res["dbmailrecipient"])


                body_part = MIMEText(MESSAGE_BODY, 'html')

                #msg['Subject'] = "Bot Status for " + date.today().strftime('%d/%m/%Y')
                msg['Subject'] = Botstatus_var + "-" + botno + "-" + Subprocess_var + "-" + Botname
                msg['From'] = "automation@redingtongroup.com"
                msg['To'] = ', '.join(to_addr)
                msg['Cc'] = ', '.join(cc_addr)
                recipients = to_addr + cc_addr
                msg.attach(body_part)

                try:
                    with open(botdir + botno + ".zip", 'rb') as file:
                        # Attach the file with filename to the email
                        msg.attach(MIMEApplication(file.read(), Name=botno + ".zip"))
                except Exception as e:
                    print(e)

                server = smtplib.SMTP("smtp.office365.com", 587)
                server.starttls()
                try:
                    server.login("automation@redingtongroup.com", "!Redb0t23#")
                    server.sendmail(msg['From'], recipients, msg.as_string())
                    server.quit()
                except Exception as e:
                    return HttpResponse(e)

            else:
                botobj.Mailsend = 0

            botobj.save()

            try:
                myfile = request.FILES['myfile']
                fs = FileSystemStorage()
                folder_name = botno
                folder_path = os.path.join(botdir, folder_name)
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                file = fs.save(os.path.join(folder_name, myfile.name), myfile)
                fileurl = fs.url(file)
            except:
                pass

            return redirect('bot:index')
    else:
        return render(request, 'bot/login.html')

def file_download(request,param):

    Botobj = TicketTrackingTable.objects.filter(projectno=param).values()
    for bot in Botobj:
        botname = bot["projectname"]
    # folder_path = "D:/COE IMPROVEMENTS/" + process + "/" + subprocess +"/" + botname
    folder_path = "E:/Botomation/data/storage1/" + str(param)
    try:
        file_list = os.listdir(folder_path)
        response = HttpResponse(content_type='application/zip')
        zip_file = zipfile.ZipFile(response, 'w')
        for file_name in file_list:
            file_path = os.path.join(folder_path, file_name)
            zip_file.write(file_path, file_name)
        zip_file.close()
        response['Content-Disposition'] = f'attachment; filename=' + botname + '.zip'
        return response
    except:
        request.session['botcreateerror'] = "FileNotFound"
        return redirect('bot:index')



def file_getfilenamelist(request,param):

    db = sqlite()
    process = db.process()
    subprocess = db.subprocess()
    botstatus = db.botstatus()
    kaizenstatus = db.kaizenstatus()

    Botobj = TicketTrackingTable.objects.filter(projectno=param).values()
    for bot in Botobj:
        processurl = bot["process"]
        subprocessurl = bot["subprocess"]
        botname = bot["projectname"]
        botno = bot["projectno"]

    # path = "D:/COE IMPROVEMENTS/" + processurl + "/" + subprocessurl +"/" + botname
    path = "E:/Botomation/data/storage1/" + str(param)
    files = []
    folders = []

    try:
        for f in os.listdir(path):
            if os.path.isfile(os.path.join(path, f)):
                files.append(f)
            else:
                folders.append(f)
    except:
        pass

   # context = {'botno': botno, 'botname': botname, 'files': files, 'folders': folders, 'show_modal': True,
   #            'process': process, 'subprocess': subprocess,
   #            'botstatus': botstatus, 'kaizenstatus': kaizenstatus }

    Bots = TicketTrackingTable.objects.filter(
        Q(status='WIP') | Q(status='Yet to Start'), projecttype='Automation').order_by('-projectno').values()

    context = {'Bots':Bots, 'process': process, 'subprocess': subprocess, 'botno': botno, 'botname': botname, 'files': files, 'folders': folders, 'show_modal': True}

    return render(request, 'bot/index.html', context)



#============================Bot View========================================

def botviewquery(request):

    session_var="allow"

    if session_var=="allow":
    #if request.session["loginvar"] == "allow":
        try:
            from_date = request.POST["fromdateview"]
            to_date = request.POST["todateview"]
        except:
            from_date = ""
            to_date = ""

        if from_date == "":
            pass
        else:
            from_date = (datetime.datetime.strptime(str(from_date), "%d/%m/%Y").strftime("%Y-%m-%d"))

        if to_date == "":
            pass
        else:
            to_date = (datetime.datetime.strptime(str(to_date), "%d/%m/%Y").strftime("%Y-%m-%d"))

        if from_date == "":
            selected_options = request.POST.getlist('dropdownselect')
            if selected_options:
                if "All" in selected_options:
                    Bots = TicketTrackingTable.objects.filter(projecttype='Automation').order_by('projectno')
                else:
                    Bots = TicketTrackingTable.objects.filter(status__in=selected_options,projecttype='Automation').order_by('projectno')
            else:

                Bots = TicketTrackingTable.objects.filter(projecttype='Automation').order_by('-projectno')

        else:
            selected_options = request.POST.getlist('dropdownselect')
            if selected_options:
                if "All" in selected_options:
                    Bots = TicketTrackingTable.objects.filter(
                        Q(startdate__range=(from_date, to_date)) | Q(enddate__range=(from_date, to_date)) | Q(
                            enhancestartdate__range=(from_date, to_date)) | Q(
                            enhanceenddate__range=(from_date, to_date)),projecttype='Automation').order_by('projectno')
                else:
                    Bots = TicketTrackingTable.objects.filter(
                        Q(startdate__range=(from_date, to_date)) | Q(enddate__range=(from_date, to_date)) | Q(
                            enhancestartdate__range=(from_date, to_date)) | Q(
                            enhanceenddate__range=(from_date, to_date)), projecttype='Automation', status__in=selected_options).order_by(
                        'projectno')
            else:
                Bots = TicketTrackingTable.objects.filter(
                    Q(startdate__range=(from_date, to_date)) | Q(enddate__range=(from_date, to_date)) | Q(
                        enhancestartdate__range=(from_date, to_date)) | Q(enhanceenddate__range=(from_date, to_date)),
                    projecttype='Automation').order_by('projectno')

        for obj in Bots:
            if "h" in str(obj.totaltime):
                manualtimespend = str(obj.manualtime)
                automationtimespend = str(obj.automationtime)
                Totaltimesaved = int(manualtimespend) - int(automationtimespend)
                Totaltimesaved = str(Totaltimesaved) + " mins"
                obj.totaltime = Totaltimesaved

        for obj in Bots:
            if "h" in str(obj.totaltime):
                print(obj.totaltime)

        for obj in Bots:

            if str(obj.startdate) != "None":
                obj.startdate = str(obj.startdate).split(" ")[0]
                obj.startdate = datetime.datetime.strptime(obj.startdate, "%Y-%m-%d")
                obj.startdate = obj.startdate.strftime("%b %d, %Y")

            if str(obj.enddate) != "None":
                obj.enddate = str(obj.enddate).split(" ")[0]
                obj.enddate = datetime.datetime.strptime(obj.enddate, "%Y-%m-%d")
                obj.enddate = obj.enddate.strftime("%b %d, %Y")

            if str(obj.enhancestartdate) != "None":
                obj.enhancestartdate = str(obj.enhancestartdate).split(" ")[0]
                obj.enhancestartdate = datetime.datetime.strptime(obj.enhancestartdate, "%Y-%m-%d")
                obj.enhancestartdate = obj.enhancestartdate.strftime("%b %d, %Y")

            if str(obj.enhanceenddate) != "None":
                obj.enhanceenddate  = str(obj.enhanceenddate ).split(" ")[0]
                obj.enhanceenddate  = datetime.datetime.strptime(obj.enhanceenddate , "%Y-%m-%d")
                obj.enhanceenddate  = obj.enhanceenddate.strftime("%b %d, %Y")


            if str(obj.manualtime) != "None" and str(obj.manualtime) != ""   and "mins" not in str(obj.manualtime):
                obj.manualtime = obj.manualtime+" "+"mins"

            if str(obj.automationtime) != "None" and str(obj.automationtime) != ""   and "mins" not in str(obj.automationtime):
                obj.automationtime = obj.automationtime+" "+"mins"

        context = {'Bots': Bots}

        return render(request, 'bot/botview.html',context)

    else:
        return render(request, 'bot/login.html')

def botview_file_download(request,param):

    Botobj = TicketTrackingTable.objects.filter(projectno=param).values()
    for bot in Botobj:
        botname = bot["projectname"]
    # folder_path = "D:/COE IMPROVEMENTS/" + process + "/" + subprocess +"/" + botname
    folder_path = "E:/Botomation/data/storage1/" + str(param)
    try:
        file_list = os.listdir(folder_path)
        response = HttpResponse(content_type='application/zip')
        zip_file = zipfile.ZipFile(response, 'w')
        for file_name in file_list:
            file_path = os.path.join(folder_path, file_name)
            zip_file.write(file_path, file_name)
        zip_file.close()
        response['Content-Disposition'] = f'attachment; filename=' + botname + '.zip'
        return response
    except:
        request.session['botcreateerror'] = "FileNotFound"
        return redirect('bot:index')


def botview_file_getfilenamelist(request,param):

    db = sqlite()
    process = db.process()
    subprocess = db.subprocess()
    botstatus = db.botstatus()
    kaizenstatus = db.kaizenstatus()

    Botobj = TicketTrackingTable.objects.filter(projectno=param).values()
    for bot in Botobj:
        processurl = bot["process"]
        subprocessurl = bot["subprocess"]
        botname = bot["projectname"]
        botno = bot["projectno"]

    # path = "D:/COE IMPROVEMENTS/" + processurl + "/" + subprocessurl +"/" + botname
    path = "E:/Botomation/data/storage1/" + str(param)
    files = []
    folders = []

    try:
        for f in os.listdir(path):
            if os.path.isfile(os.path.join(path, f)):
                files.append(f)
            else:
                folders.append(f)
    except:
        pass

    # context = {'botno': botno, 'botname': botname, 'files': files, 'folders': folders, 'show_modal': True,
    #            'process': process, 'subprocess': subprocess,
    #            'botstatus': botstatus, 'kaizenstatus': kaizenstatus }

    Bots = TicketTrackingTable.objects.all().values().order_by('-projectno')

    for bot in Bots:
        print(bot["projectno"])

    context = {'Bots': Bots, 'process': process, 'subprocess': subprocess, 'botno': botno, 'botname': botname,
               'files': files, 'folders': folders, 'show_modal': True}

    return render(request, 'bot/botview.html', context)


def downloadreport(request):

    session_var = "allow"

    if session_var == "allow":

        from_date = request.POST["fromdatenew"]
        to_date = request.POST["todatenew"]

        if from_date=="":
            pass
        else:
            from_date=(datetime.datetime.strptime(str(from_date),"%d/%m/%Y").strftime("%Y-%m-%d"))

        if to_date == "":
            pass
        else:
            to_date=(datetime.datetime.strptime(str(to_date), "%d/%m/%Y").strftime("%Y-%m-%d"))

        selected_options = list(request.POST.getlist('hiddenInput'))

        if from_date=="":

           my_queryset = TicketTrackingTable.objects.filter(projecttype='Automation').order_by('projectno')

        else:

            if len(selected_options[0])>0:

                if "New" in selected_options[0] and "Enhancement" in selected_options[0]:
                    my_queryset = TicketTrackingTable.objects.filter(
                        Q(startdate__range=(from_date, to_date)) | Q(enddate__range=(from_date, to_date)) | Q(
                            enhancestartdate__range=(from_date, to_date)) | Q(
                            enhanceenddate__range=(from_date, to_date)), projecttype='Automation').order_by(
                        'projectno')
                else:
                    if "New" in selected_options[0]:
                        my_queryset = TicketTrackingTable.objects.filter(
                            Q(startdate__range=(from_date, to_date)) | Q(enddate__range=(from_date, to_date)), projecttype='Automation').order_by('projectno')
                    else:
                        my_queryset = TicketTrackingTable.objects.filter(
                             Q(enhancestartdate__range=(from_date, to_date)) | Q(enhanceenddate__range=(from_date, to_date)), projecttype='Automation').order_by('projectno')

            else:

                my_queryset = TicketTrackingTable.objects.filter(
                    Q(startdate__range=(from_date, to_date)) | Q(enddate__range=(from_date, to_date)) | Q(
                        enhancestartdate__range=(from_date, to_date)) | Q(
                        enhanceenddate__range=(from_date, to_date)),projecttype='Automation').order_by('projectno')

        wb = Workbook()

        ws = wb.active

        today = date.today()

        #ws.append(['Botno-1', 'Botname-2','Botdesc-20','Process-3','Subprocess-4','Spocname-5','RequestBy-6','Teamlead-7','DevelopmentBy-8','Technologyused-9','Creationdate-10','Startdate-11','Enddate-12','Botstatus-19','Manualtimespend-13','Automationtimespend-14','Totaltimesaved-15','Totaldaysaved-16','Kaizenawardstatus-17','Kaizenawardedyear-18'])
        ws.append(['Projectno','Projectname','Business Unit','Process','Subprocess','Spocname','RequestBy','Teamlead', 'ManagerMail' ,'DevelopmentBy','Technologyused','Creationdate','Startdate','Enddate','EnhancementStartdate','EnhancementEnddate','Manualtimespend','Automationtimespend','Totaltimesaved','Totaldaysaved','Kaizenawardstatus','Projectstatus','Livestatus','Categorization','Projectdesc','Comments','Requesttype','Reason'])

        for obj in my_queryset:

            if str(obj.startdate)!="None":
                print(str(obj.startdate))
                startdate = datetime.datetime.strptime(str(obj.startdate),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
            else:
                startdate=""

            if str(obj.creationdate)!="None":
                creationdate = datetime.datetime.strptime(str(obj.creationdate),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
            else:
                creationdate=""

            if str(obj.enddate)!="None":
                enddate = datetime.datetime.strptime(str(obj.enddate),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
            else:
                enddate=""

            if str(obj.enhancestartdate)!="None" and "1900" not in str(obj.enhancestartdate):
                enhancestartdate = datetime.datetime.strptime(str(obj.enhancestartdate),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
            else:
                enhancestartdate=""

            if str(obj.enhanceenddate)!="None" and "1900" not in str(obj.enhanceenddate):
                enhanceenddate = datetime.datetime.strptime(str(obj.enhanceenddate),"%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
            else:
                enhanceenddate=""

            if "mins" in str(obj.totalday) and "days" not in str(obj.totalday):
                totaldaysaved = round(int(str(obj.totalday).replace("mins", "")) / 60)
                if totaldaysaved <= 8:
                    totaldaysaved = str(totaldaysaved) + " hrs"
                else:
                    totaldaysaved = str(round(totaldaysaved / 8)) + " days"
            else:
                totaldaysaved = obj.totalday

            if str(totaldaysaved) == "None":
                totaldaysaved = ""

            if str(obj.manualtime) == "None":
                manualtimespend = ""
            else:
                manualtimespend = str(obj.manualtime)

            if str(obj.automationtime) == "None":
                automationtimespend = ""
            else:
                automationtimespend = str(obj.automationtime)

            if str(obj.totaltime) == "None":
                Totaltimesaved = ""
            elif ":" in str(obj.totaltime):
                print(str(obj.totaltime))
                Totaltimesaved = str(obj.totaltime)
                hour = Totaltimesaved.split(":")[0]
                minute = Totaltimesaved.split(":")[1]
                hourinmins = int(hour)*60
                Totaltimesaved = str(hourinmins+int(minute))+str(" mins")
            elif "h" in str(obj.totaltime):
                Totaltimesaved = int(manualtimespend)-int(automationtimespend)
                Totaltimesaved = str(Totaltimesaved)+ " mins"
            else:
                Totaltimesaved = str(obj.totaltime)


            ws.append([obj.projectno, obj.projectname, obj.businessunit, obj.process, obj.subprocess, obj.spocname, obj.requestormailid, obj.teamleadmailid, obj.managermailid,
                 obj.developermailid, obj.technologyused, creationdate, startdate, enddate, enhancestartdate, enhanceenddate,
                 manualtimespend, automationtimespend, Totaltimesaved, totaldaysaved, obj.kaizenstatus,
                 obj.status,obj.livestatus,obj.categorization,obj.projectdesc,obj.remarks,obj.request_type,obj.reason])

        # Generate the response
        '''
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 0.5
            ws.column_dimensions[column].width = adjusted_width
        '''

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="BotDetails_{}.xlsx"'.format(today.strftime("%d-%m-%Y"))
        wb.save(response)

        return response
    else:
        return render(request, 'bot/login.html')


#=============================History=====================================


def history(request, param):
    bots_hist = BotHist.objects.filter(botno=param)

    BotHistoryobj = BotHist.objects.all().values()
    bot_hist = set()
    for bot in BotHistoryobj:
        bot_hist.add(bot["botno"])

    bot = Bot.objects.filter(Botno=param)
    #Bots = Bot.objects.all().order_by('-Botno')
    Bots = Bot.objects.filter(Q(Botstatus='Devolopment In Progress') | Q(Botstatus='Bot No.Generated')).order_by('-Botno').values()
    newbotno = Bot.objects.all().aggregate(Max('Botno'))
    for botobj in bots_hist:
        botobj.remarks=str(botobj.remarks).replace("None","")
        botobj.save()
        botnamevar = botobj.botname

    context = {'bothistory':bot_hist,'newbotno': newbotno['Botno__max'] + 1,'show_hist_modal': True, 'bots_hist':bots_hist,'Bot': bot,'Bots': Bots,'Botno':param, 'Botname': botnamevar }
    return render(request, 'bot/index.html', context)


def history_viewpage(request, param):

    bots_hist = BotHist.objects.filter(botno=param)

    BotHistoryobj = BotHist.objects.all().values()
    bot_hist = set()
    for bot in BotHistoryobj:
        bot_hist.add(bot["botno"])

    bot = Bot.objects.filter(Botno=param)
    Bots = Bot.objects.all().order_by('-Botno')
    #Bots = Bot.objects.filter(Q(Botstatus='Devolopment In Progress') | Q(Botstatus='Bot No.Generated')).order_by('-Botno').values()
    newbotno = Bot.objects.all().aggregate(Max('Botno'))
    for botobj in bots_hist:
        botobj.remarks=str(botobj.remarks).replace("None","")
        botobj.save()
        botnamevar = botobj.botname
    context = {'bothistory':bot_hist,'newbotno': newbotno['Botno__max'] + 1,'show_hist_modal': True, 'bots_hist':bots_hist,'Bot': bot,'Bots': Bots,'Botno':param, 'Botname': botnamevar }
    return render(request, 'bot/botview.html', context)



#===========================Chart View====================================

def sample_bar_chart(request):

    if request.session["loginvar"] == "allow":

        Totalcount = Bot.objects.all().count()
        Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated').count()
        devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress').count()
        usertesting = Bot.objects.filter(Botstatus='Under User Testing').count()
        #usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
        hold = Bot.objects.filter(Botstatus='Hold').count()
        tm = Bot.objects.filter(Botstatus='TM TO BE DONE').count()
        completed = Bot.objects.filter(Botstatus='Completed').count()
        cancelled = Bot.objects.filter(Botstatus='Cancelled').count()

        datapoints = [
            {"y": Botnogenerated, "label": "BotCreated","indexLabelPlacement": "outside"},
            {"y": devlopment_in_progress , "label": "Devolopment In Progress", "indexLabelPlacement": "outside"},
            {"y": usertesting, "label": "UserTesting","indexLabelPlacement": "outside"},
            {"y": hold, "label": "Hold", "indexLabelPlacement": "outside"},
            {"y": tm, "label": "TimeStudy", "indexLabelPlacement": "outside"},
            {"y": completed, "label": "Completed", "indexLabelPlacement": "outside"},
            {"y": cancelled, "label": "Cancelled", "indexLabelPlacement": "outside"},
        ]

        bot_data=[]

        return render(request, 'bot/chart.html', {"datapoints1": datapoints,"datapoints2": datapoints,"datapoints3": datapoints,"datapoints4": datapoints,"bot_data" : bot_data,"chartview":"bar", "Totalcount":Totalcount})
    else:
        return render(request, 'bot/login.html')

def changeChart(request):

    if request.session["loginvar"] == "allow":

        status_list = []
        status_list.append('Bot No.Generated')
        status_list.append('Devolopment In Progress')
        status_list.append('Under User Testing')
        status_list.append('TM TO BE DONE')
        status_list.append('Completed')

        selected_options = request.POST["hiddenInput"]
        charttype = request.POST["chartselect"]

        if selected_options=="All":
            Totalcount = Bot.objects.all().count()
            Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated').count()
            devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress').count()
            usertesting = Bot.objects.filter(Botstatus='Under User Testing').count()

            hold = Bot.objects.filter(Botstatus='Hold').count()
            tm = Bot.objects.filter(Botstatus='TM TO BE DONE').count()
            completed = Bot.objects.filter(Botstatus='Completed').count()
            cancelled = Bot.objects.filter(Botstatus='Cancelled').count()

        elif selected_options == "Businessunit":
            rgf_Totalcount = Bot.objects.filter(businessunit='RGF').count()
            rgf_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated',businessunit='RGF').count()
            rgf_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',businessunit='RGF').count()
            rgf_usertesting = Bot.objects.filter(Botstatus='Under User Testing',businessunit='RGF').count()
            rgf_tm = Bot.objects.filter(Botstatus='TM TO BE DONE',businessunit='RGF').count()
            rgf_completed = Bot.objects.filter(Botstatus='Completed',businessunit='RGF').count()

            rgs_Totalcount = Bot.objects.filter(businessunit='RGS').count()
            rgs_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated',businessunit='RGS').count()
            rgs_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',businessunit='RGS').count()
            rgs_usertesting = Bot.objects.filter(Botstatus='Under User Testing',businessunit='RGS').count()
            rgs_tm = Bot.objects.filter(Botstatus='TM TO BE DONE',businessunit='RGS').count()
            rgs_completed = Bot.objects.filter(Botstatus='Completed',businessunit='RGS').count()

            wabco_Totalcount = Bot.objects.filter(businessunit='Wabco').count()
            wabco_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', businessunit='Wabco').count()
            wabco_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',businessunit='Wabco').count()
            wabco_usertesting = Bot.objects.filter(Botstatus='Under User Testing', businessunit='Wabco').count()
            wabco_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', businessunit='Wabco').count()
            wabco_completed = Bot.objects.filter(Botstatus='Completed', businessunit='Wabco').count()


        elif selected_options=="Team":
            Prajnya_Totalcount = Bot.objects.filter(Developermail="prajnya.sahu@redingtongroup.com", Botstatus__in=status_list).count()
            Prajnya_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', Developermail="prajnya.sahu@redingtongroup.com").count()
            Prajnya_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',Developermail="prajnya.sahu@redingtongroup.com").count()
            Prajnya_usertesting = Bot.objects.filter(Botstatus='Under User Testing', Developermail="prajnya.sahu@redingtongroup.com").count()
            # usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
            Prajnya_hold = Bot.objects.filter(Botstatus='Hold', Developermail="prajnya.sahu@redingtongroup.com").count()
            Prajnya_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', Developermail="prajnya.sahu@redingtongroup.com").count()
            Prajnya_completed = Bot.objects.filter(Botstatus='Completed', Developermail="prajnya.sahu@redingtongroup.com").count()
            Prajnya_cancelled = Bot.objects.filter(Botstatus='Cancelled', Developermail="prajnya.sahu@redingtongroup.com").count()

            Prashanth_Totalcount = Bot.objects.filter(Developermail="prasanth.muruga@redingtongroup.com", Botstatus__in=status_list).count()
            Prashanth_Botnogenerated = Bot.objects.filter(Botstatus="prasanth.muruga@redingtongroup.com", Developermail=selected_options).count()
            Prashanth_devlopment_in_progress = Bot.objects.filter(Botstatus="prasanth.muruga@redingtongroup.com",Developermail=selected_options).count()
            Prashanth_usertesting = Bot.objects.filter(Botstatus="prasanth.muruga@redingtongroup.com", Developermail=selected_options).count()
            # usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
            Prashanth_hold = Bot.objects.filter(Botstatus='Hold', Developermail="prasanth.muruga@redingtongroup.com").count()
            Prashanth_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', Developermail="prasanth.muruga@redingtongroup.com").count()
            Prashanth_completed = Bot.objects.filter(Botstatus='Completed', Developermail="prasanth.muruga@redingtongroup.com").count()
            Prashanth_cancelled = Bot.objects.filter(Botstatus='Cancelled', Developermail="prasanth.muruga@redingtongroup.com").count()

            Santhosh_Totalcount = Bot.objects.filter(Developermail="santhosh.arjunan@redingtongroup.com", Botstatus__in=status_list).count()
            Santhosh_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', Developermail="santhosh.arjunan@redingtongroup.com").count()
            Santhosh_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',Developermail="santhosh.arjunan@redingtongroup.com").count()
            Santhosh_usertesting = Bot.objects.filter(Botstatus='Under User Testing', Developermail="santhosh.arjunan@redingtongroup.com").count()
            # usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
            Santhosh_hold = Bot.objects.filter(Botstatus='Hold', Developermail="santhosh.arjunan@redingtongroup.com").count()
            Santhosh_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', Developermail="santhosh.arjunan@redingtongroup.com").count()
            Santhosh_completed = Bot.objects.filter(Botstatus='Completed', Developermail="santhosh.arjunan@redingtongroup.com").count()
            Santhosh_cancelled = Bot.objects.filter(Botstatus='Cancelled', Developermail="santhosh.arjunan@redingtongroup.com").count()

            Thanis_Totalcount = Bot.objects.filter(Developermail="thanis.albert@redingtongroup.com", Botstatus__in=status_list).count()
            Thanis_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', Developermail="thanis.albert@redingtongroup.com").count()
            Thanis_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',Developermail="thanis.albert@redingtongroup.com").count()
            Thanis_usertesting = Bot.objects.filter(Botstatus='Under User Testing', Developermail="thanis.albert@redingtongroup.com").count()
            # usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
            Thanis_hold = Bot.objects.filter(Botstatus='Hold', Developermail="thanis.albert@redingtongroup.com").count()
            Thanis_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', Developermail="thanis.albert@redingtongroup.com").count()
            Thanis_completed = Bot.objects.filter(Botstatus='Completed', Developermail="thanis.albert@redingtongroup.com").count()
            Thanis_cancelled = Bot.objects.filter(Botstatus='Cancelled', Developermail="thanis.albert@redingtongroup.com").count()

        else:
            Totalcount = Bot.objects.filter(Developermail=selected_options, Botstatus__in=status_list).count()
            Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', Developermail=selected_options).count()
            devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress',Developermail=selected_options).count()
            usertesting = Bot.objects.filter(Botstatus='Under User Testing', Developermail=selected_options).count()
            # usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
            hold = Bot.objects.filter(Botstatus='Hold', Developermail=selected_options).count()
            tm = Bot.objects.filter(Botstatus='TM TO BE DONE', Developermail=selected_options).count()
            completed = Bot.objects.filter(Botstatus='Completed', Developermail=selected_options).count()
            cancelled = Bot.objects.filter(Botstatus='Cancelled', Developermail=selected_options).count()


        if (charttype =="bar" or charttype=="pie") and selected_options !="Team" and selected_options != "Businessunit":

            datapoints = []
            if Botnogenerated > 0:
                datapoints.append({"y": Botnogenerated, "label": "BotCreated", "indexLabelPlacement": "outside"})
            if devlopment_in_progress > 0:
                datapoints.append(
                    {"y": devlopment_in_progress, "label": "devlopment_in_progress", "indexLabelPlacement": "outside"})
            if usertesting > 0:
                datapoints.append({"y": usertesting, "label": "UserTesting", "indexLabelPlacement": "outside"})
            if hold > 0 and selected_options == "All":
                datapoints.append({"y": hold, "label": "Hold", "indexLabelPlacement": "outside"})
            if cancelled > 0 and selected_options == "All":
                datapoints.append({"y": cancelled, "label": "Cancelled", "indexLabelPlacement": "outside"})
            if tm > 0:
                datapoints.append({"y": tm, "label": "TimeStudy", "indexLabelPlacement": "outside"})
            if completed > 0:
                datapoints.append({"y": completed, "label": "Completed", "indexLabelPlacement": "outside"})

            bot_data = []
            if Botnogenerated > 0:
                bot_data.append({"label": "BotCreated", "y": Botnogenerated})
            if devlopment_in_progress > 0:
                bot_data.append({"label": "Devolopment In Progress", "y": devlopment_in_progress})
            if usertesting > 0:
                bot_data.append({"label": "UserTesting", "y": usertesting})
            if hold > 0 and selected_options == "All":
                bot_data.append({"label": "Hold", "y": hold})
            if cancelled > 0 and selected_options == "All":
                bot_data.append({"label": "Cancelled", "y": cancelled})
            if tm > 0:
                bot_data.append({"label": "TimeStudy", "y": tm})
            if completed > 0:
                bot_data.append({"label": "completed", "y": completed})

        if (charttype == "bar" or charttype == "pie") and selected_options == "Businessunit":

            Totalcount = rgs_Totalcount + rgf_Totalcount + wabco_Totalcount;

            rgs_datapoints = [
                {"label": "BotCreated", "y": rgs_Botnogenerated, "indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": rgs_devlopment_in_progress, "indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": rgs_usertesting, "indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": rgs_tm, "indexLabelPlacement": "outside"},
                {"label": "completed", "y": rgs_completed, "indexLabelPlacement": "outside"}
            ]

            rgf_datapoints = [
                {"label": "BotCreated", "y": rgf_Botnogenerated, "indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": rgf_devlopment_in_progress,
                 "indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": rgf_usertesting, "indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": rgf_tm, "indexLabelPlacement": "outside"},
                {"label": "completed", "y": rgf_completed, "indexLabelPlacement": "outside"},
            ]

            wabco_datapoints = [
                {"label": "BotCreated", "y": wabco_Botnogenerated, "indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": wabco_devlopment_in_progress,
                 "indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": wabco_usertesting, "indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": wabco_tm, "indexLabelPlacement": "outside"},
                {"label": "completed", "y": wabco_completed, "indexLabelPlacement": "outside"},
            ]

            bot_data = []
            return render(request, 'bot/chart.html',
                          {"datapoints1": rgs_datapoints, "datapoints2": rgf_datapoints,
                           "datapoints3": wabco_datapoints, "datapoints4": rgf_datapoints, "bot_data": bot_data,
                           "chartview": "businessunitbar", "Totalcount": Totalcount, "Rgscount": rgs_Totalcount,
                           "Rgfcount": rgf_Totalcount,"Wabcocount": wabco_Totalcount, "Developername": selected_options})

        if (charttype =="bar" or charttype=="pie") and selected_options =="Team":

            TeamTotalcount=Prashanth_Totalcount+Prajnya_Totalcount+Santhosh_Totalcount+Thanis_Totalcount;

            Prajnya_datapoints = [
                {"label": "BotCreated", "y": Prajnya_Botnogenerated,"indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": Prajnya_devlopment_in_progress,"indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": Prajnya_usertesting,"indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": Prajnya_tm,"indexLabelPlacement": "outside"},
                {"label": "completed", "y": Prajnya_completed,"indexLabelPlacement": "outside"}
            ]

            Prashanth_datapoints = [
                {"label": "BotCreated", "y": Prashanth_Botnogenerated,"indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": Prashanth_devlopment_in_progress,"indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": Prashanth_usertesting,"indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": Prashanth_tm,"indexLabelPlacement": "outside"},
                {"label": "completed", "y": Prashanth_completed,"indexLabelPlacement": "outside"},
            ]

            Santhosh_datapoints = [
                {"label": "BotCreated", "y": Santhosh_Botnogenerated,"indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": Santhosh_devlopment_in_progress,"indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": Santhosh_usertesting,"indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": Santhosh_tm,"indexLabelPlacement": "outside"},
                {"label": "completed", "y": Santhosh_completed,"indexLabelPlacement": "outside"},
            ]

            Thanis_datapoints = [
                {"label": "BotCreated", "y": Thanis_Botnogenerated,"indexLabelPlacement": "outside"},
                {"label": "Devolopment In Progress", "y": Thanis_devlopment_in_progress,"indexLabelPlacement": "outside"},
                {"label": "UserTesting", "y": Thanis_usertesting,"indexLabelPlacement": "outside"},
                {"label": "TimeStudy", "y": Thanis_tm,"indexLabelPlacement": "outside"},
                {"label": "completed", "y": Thanis_completed,"indexLabelPlacement": "outside"},
            ]

            bot_data=[]
            return render(request, 'bot/chart.html',
                          {"datapoints1":Prajnya_datapoints,"datapoints2":Prashanth_datapoints, "datapoints3":Santhosh_datapoints, "datapoints4":Thanis_datapoints, "bot_data": bot_data, "chartview": "teambar", "Totalcount": TeamTotalcount,"Prajnyacount":Prajnya_Totalcount,"Prasanthcount":Prashanth_Totalcount, "Santhoshcount": Santhosh_Totalcount, "Thaniscount": Thanis_Totalcount,
                           "Developername": selected_options})

        if charttype == "pie" and selected_options != "Businessunit" and selected_options != "Team" :
            return render(request, 'bot/chart.html',{"datapoints1": datapoints, "datapoints2": datapoints,"datapoints3": datapoints, "datapoints4": datapoints,"bot_data": bot_data, "chartview": "pie", "Totalcount": Totalcount, "Developername":selected_options})

        if charttype == "bar" and selected_options != "Businessunit" and selected_options != "Team" :
            return render(request, 'bot/chart.html',{"datapoints1": datapoints,"datapoints2": datapoints,"datapoints3": datapoints,"datapoints4": datapoints, "bot_data": bot_data, "chartview": "bar", "Totalcount": Totalcount, "Developername":selected_options})

        if charttype == "select":
            return HttpResponse("select")
    else:
        return render(request, 'bot/login.html')


def chart_mail2(request):
    rgf_Totalcount = Bot.objects.filter(businessunit='RGF').count()
    rgf_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', businessunit='RGF').count()
    rgf_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress', businessunit='RGF').count()
    rgf_usertesting = Bot.objects.filter(Botstatus='Under User Testing', businessunit='RGF').count()
    rgf_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', businessunit='RGF').count()
    rgf_completed = Bot.objects.filter(Botstatus='Completed', businessunit='RGF').count()

    rgs_Totalcount = Bot.objects.filter(businessunit='RGS').count()
    rgs_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', businessunit='RGS').count()
    rgs_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress', businessunit='RGS').count()
    rgs_usertesting = Bot.objects.filter(Botstatus='Under User Testing', businessunit='RGS').count()
    rgs_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', businessunit='RGS').count()
    rgs_completed = Bot.objects.filter(Botstatus='Completed', businessunit='RGS').count()

    wabco_Totalcount = Bot.objects.filter(businessunit='Wabco').count()
    wabco_Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated', businessunit='Wabco').count()
    wabco_devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress', businessunit='Wabco').count()
    wabco_usertesting = Bot.objects.filter(Botstatus='Under User Testing', businessunit='Wabco').count()
    wabco_tm = Bot.objects.filter(Botstatus='TM TO BE DONE', businessunit='Wabco').count()
    wabco_completed = Bot.objects.filter(Botstatus='Completed', businessunit='Wabco').count()



    Totalcount = rgs_Totalcount + rgf_Totalcount+wabco_Totalcount;
    rgs_datapoints = [
        {"label": "BotCreated", "y": rgs_Botnogenerated, "indexLabelPlacement": "outside"},
        {"label": "Devolopment In Progress", "y": rgs_devlopment_in_progress, "indexLabelPlacement": "outside"},
        {"label": "UserTesting", "y": rgs_usertesting, "indexLabelPlacement": "outside"},
        {"label": "TimeStudy", "y": rgs_tm, "indexLabelPlacement": "outside"},
        {"label": "completed", "y": rgs_completed, "indexLabelPlacement": "outside"}
    ]

    rgf_datapoints = [
        {"label": "BotCreated", "y": rgf_Botnogenerated, "indexLabelPlacement": "outside"},
        {"label": "Devolopment In Progress", "y": rgf_devlopment_in_progress,
         "indexLabelPlacement": "outside"},
        {"label": "UserTesting", "y": rgf_usertesting, "indexLabelPlacement": "outside"},
        {"label": "TimeStudy", "y": rgf_tm, "indexLabelPlacement": "outside"},
        {"label": "completed", "y": rgf_completed, "indexLabelPlacement": "outside"},
    ]

    wabco_datapoints = [
        {"label": "BotCreated", "y": wabco_Botnogenerated, "indexLabelPlacement": "outside"},
        {"label": "Devolopment In Progress", "y": wabco_devlopment_in_progress,
         "indexLabelPlacement": "outside"},
        {"label": "UserTesting", "y": wabco_usertesting, "indexLabelPlacement": "outside"},
        {"label": "TimeStudy", "y": wabco_tm, "indexLabelPlacement": "outside"},
        {"label": "completed", "y": wabco_completed, "indexLabelPlacement": "outside"},
    ]


    bot_data = []
    return render(request, 'bot/mailchart2.html',
                  {"datapoints1": rgs_datapoints, "datapoints2": rgf_datapoints,
                   "datapoints3": wabco_datapoints, "datapoints4": rgf_datapoints, "bot_data": bot_data,
                   "chartview": "businessunitbar", "Totalcount": Totalcount, "Rgscount": rgs_Totalcount,
                   "Rgfcount": rgf_Totalcount, "Wabcocount":wabco_Totalcount})


def chart_mail(request):
    Totalcount = Bot.objects.all().count()
    Botnogenerated = Bot.objects.filter(Botstatus='Bot No.Generated').count()
    devlopment_in_progress = Bot.objects.filter(Botstatus='Devolopment In Progress').count()
    usertesting = Bot.objects.filter(Botstatus='Under User Testing').count()
    # usertestanddev = Bot.objects.filter(Botstatus='User Testing & Dev Enhancement').count()
    hold = Bot.objects.filter(Botstatus='Hold').count()
    tm = Bot.objects.filter(Botstatus='TM TO BE DONE').count()
    completed = Bot.objects.filter(Botstatus='Completed').count()
    cancelled = Bot.objects.filter(Botstatus='Cancelled').count()

    datapoints = [
        {"y": Botnogenerated, "label": "BotCreated", "indexLabelPlacement": "outside"},
        {"y": devlopment_in_progress, "label": "Devolopment In Progress", "indexLabelPlacement": "outside"},
        {"y": usertesting, "label": "UserTesting", "indexLabelPlacement": "outside"},
        {"y": hold, "label": "Hold", "indexLabelPlacement": "outside"},
        {"y": tm, "label": "TimeStudy", "indexLabelPlacement": "outside"},
        {"y": completed, "label": "Completed", "indexLabelPlacement": "outside"},
        {"y": cancelled, "label": "Cancelled", "indexLabelPlacement": "outside"},
    ]

    bot_data = []

    return render(request, 'bot/mailchart.html',
                  {"cancelled":cancelled, "completed": completed, "tm": tm, "hold":hold,   "usertesting":usertesting,  "devlopment_in_progress":  devlopment_in_progress, "Botgenerated": Botnogenerated, "datapoints1": datapoints, "datapoints2": datapoints, "datapoints3": datapoints,
                   "datapoints4": datapoints, "bot_data": bot_data, "chartview": "bar", "Totalcount": Totalcount})



def mailreport(request):

    if request.session["loginvar"] == "allow":

        options = Options()
        options.headless = False  # Run Chrome in headless mode

        url = "http://172.26.1.19:85/bot/mailchart.html"
        #url = "http://172.24.3.13:85/bot/chart.html"


        driver = webdriver.Chrome(executable_path="E:\\chromedriver.exe", options=options)
        #driver = webdriver.Chrome(executable_path="D:\\chromedriver.exe", options=options)

        Totalcount = Bot.objects.all().count()

        driver.get(url)

        # Wait for the chart to be generated (replace "chart-element-id" with the actual ID of the chart element)
        chart_element = driver.find_element_by_id("chartContainer")

        time.sleep(4)

        chart_image = chart_element.screenshot_as_png

        with open(settings.MEDIA_ROOT + "\\chart.png", "wb") as f:
            f.write(chart_image)


        from_date = request.POST["fromdatemailnew"]
        to_date = request.POST["todatemailnew"]

        if from_date == "":
            pass
        else:
            from_date = (datetime.datetime.strptime(str(from_date), "%d/%m/%Y").strftime("%Y-%m-%d"))

        if to_date == "":
            pass
        else:
            to_date = (datetime.datetime.strptime(str(to_date), "%d/%m/%Y").strftime("%Y-%m-%d"))


        selected_options = request.POST.getlist('hiddenInputmail')
        for select in selected_options:
            temp = str(select)

        try:
            selected_options = temp.split(",")
        except:
            selected_options = ""


        if from_date=="":

            if len(selected_options[0])>0:
                if "All" in selected_options:
                    my_queryset = Bot.objects.all().order_by('Botno')
                else:

                    my_queryset = Bot.objects.filter(Botstatus__in=selected_options).order_by('Botno')
            else:
                my_queryset = Bot.objects.all().order_by('Botno')
        else:

            if len(selected_options[0])>0:
                if "All" in selected_options:
                    my_queryset = Bot.objects.all().filter(Startdate__range=[from_date, to_date]).order_by('Botno')
                else:
                    my_queryset = Bot.objects.filter(Startdate__range=[from_date, to_date],Botstatus__in=selected_options).order_by('Botno')
            else:
                my_queryset = Bot.objects.all().filter(Startdate__range=[from_date, to_date]).order_by('Botno')


        wb = Workbook()
        ws = wb.active

        headers = ['Botno','Botname','Process','Subprocess','Spocname','RequestBy','Teamlead','DevelopmentBy','Technologyused','Creationdate','Startdate','Enddate','Enhancementstartdate','Enhancementenddate','Manualtimespend','Automationtimespend','Totaltimesaved','Totaldaysaved','Kaizenawardstatus','Kaizenawardedyear','Botstatus','Botdesc']

        ws.append(headers)

        for obj in my_queryset:
            if str(obj.Creationdate) != "None":
                creationdate = datetime.datetime.strptime(str(obj.Creationdate), "%Y-%m-%d").strftime("%d-%m-%Y")
            else:
                creationdate = ""

            if str(obj.Startdate) != "None":
                startdate = datetime.datetime.strptime(str(obj.Startdate), "%Y-%m-%d").strftime("%d-%m-%Y")
            else:
                startdate = ""

            if str(obj.Enddate) != "None":
                enddate = datetime.datetime.strptime(str(obj.Enddate), "%Y-%m-%d").strftime("%d-%m-%Y")
            else:
                enddate = ""

            if str(obj.enhancestartdate) != "None":
                enhancementstartdate = datetime.datetime.strptime(str(obj.enhancestartdate), "%Y-%m-%d").strftime("%d-%m-%Y")
            else:
                enhancementstartdate = ""


            if str(obj.enhanceenddate) != "None":
                enhancementenddate = datetime.datetime.strptime(str(obj.enhanceenddate), "%Y-%m-%d").strftime("%d-%m-%Y")
            else:
                enhancementenddate = ""

            row = [obj.Botno,obj.Botname,obj.Process,obj.Subprocess,obj.Spocname,obj.Requestormail,obj.Teamleadmail,obj.Developermail,obj.Technologyused,creationdate,startdate,enddate,enhancementstartdate,enhancementenddate,obj.Manualtimespend,obj.Automationtimespend,obj.Totaltimesaved,obj.Totaldaysaved,obj.Kaizenawardstatus,obj.Kaizenawardyear,obj.Botstatus,obj.Botdesc]

            ws.append(row)

        wb.save(settings.MEDIA_ROOT + "\\RedBot.xlsx")

        msg = MIMEMultipart('related')
        db_name = "All"
        MESSAGE_BODY ="""\
        <html>
          <head></head>
            <body>
              Dear """ + db_name + """ <br><br>
              Greetings for the day!!!<br><br>
              Please find the Bot status for """ +date.today().strftime('%d/%m/%Y')+"""<br><br>
              <br> <h4>Total Bots: """+ str(Totalcount) +"""</h4><br><img src="cid:image1"> <br><br> Thanks,<br>Team Automation 
            </body>
        </html>
        """

        body_part = MIMEText(MESSAGE_BODY, 'html')

        to_addr_queryset = Mailreport_to.objects.all().values()
        for to_addr in to_addr_queryset:
            to_addr = str(to_addr["To_address"]).split(";")

        #cc_addr_queryset = Mailreport_cc.objects.all().values()
        #for cc_addr in cc_addr_queryset:
        #    cc_addr = str(cc_addr["Cc_address"]).split(";")

        to_addr = []
        #to_addr.append("rathina.moorthy@redingtongroup.com")
        to_addr.append("thanis.albert@redingtongroup.com")
        #cc_addr = []
        #cc_addr.append("thanis.a@redingtongroup.com")

        #recipients = to_addr + cc_addr
        #recipients = 'rathina.moorthy@redingtongroup.com,thanis.albert@redingtongroup.com'
        #recipients = 'thanis.albert@redingtongroup.com'

        recipients = to_addr;

        msg['Subject'] = "Bot Status for "+date.today().strftime('%d/%m/%Y')
        msg['From'] = "automation@redingtongroup.com"
        msg['To'] = ', '.join(to_addr)
        #msg['Cc'] = ', '.join(cc_addr)
        msg.attach(body_part)

        with open(settings.MEDIA_ROOT +"\\chart.png", 'rb') as f:
            img_data = f.read()
        img = MIMEImage(img_data)
        img.add_header('Content-ID', '<image1>')
        msg.attach(img)

        with open(settings.MEDIA_ROOT +"\\RedBot.xlsx", 'rb') as file:
            msg.attach(MIMEApplication(file.read(), Name="RedBotReport.xlsx"))

        server = smtplib.SMTP("smtp.office365.com", 587)
        server.starttls()
        try:
            server.login("automation@redingtongroup.com", "!Redb0t23#")
            server.sendmail(msg['From'], recipients, msg.as_string())
            server.quit()
            request.session["mailsent"]="sent"
            return redirect('bot:botviewquery')
        except Exception as e:
            return HttpResponse(e)
    else:
        return render(request, 'bot/login.html')


















